from __future__ import annotations

import asyncio
import base64
import contextlib
import csv
import html as _html
import importlib
import io
import json
import mimetypes
import os
import re
import tarfile
import uuid
import zipfile
from collections.abc import AsyncIterator, Awaitable, Callable, Sequence
from datetime import date, datetime, timezone
from pathlib import Path, PurePosixPath
from typing import Any, cast
from urllib.parse import parse_qsl, quote, urlencode, urlsplit, urlunsplit

import httpx
from fastapi import APIRouter, Depends, File, Header, HTTPException, Request, UploadFile, WebSocket, WebSocketDisconnect
from fastapi.responses import FileResponse, HTMLResponse, RedirectResponse, Response, StreamingResponse
from starlette.datastructures import UploadFile as StarletteUploadFile

from ragtime.config.settings import settings
from ragtime.core.app_settings import get_app_settings
from ragtime.core.auth import decode_access_token, get_browser_matched_origin
from ragtime.core.auth_methods import build_auth_method_statuses, normalize_auth_method_key
from ragtime.core.logging import get_logger
from ragtime.core.rate_limit import SHARE_AUTH_RATE_LIMIT, limiter
from ragtime.core.security import get_current_user, get_current_user_optional, get_session_token
from ragtime.core.user_identity import (
    USER_FINGERPRINT_SCOPE_WORKSPACE,
    build_user_fingerprint_subject,
    build_workspace_user_fingerprint,
    normalize_user_identity,
)
from ragtime.core.userspace_limits import (
    USERSPACE_PRIMITIVE_ARCHIVE_DEFAULT_MAX_ENTRIES,
    USERSPACE_PRIMITIVE_UPLOAD_DEFAULT_MAX_BYTES,
    clamp_userspace_primitive_archive_max_entries,
    clamp_userspace_primitive_upload_max_bytes,
)
from ragtime.indexer.document_parser import extract_text_from_file_async
from ragtime.userspace.html_templates import render_preview_host_unreachable_page_html
from ragtime.userspace.models import (
    UserSpaceAuthMethod,
    UserSpaceBrowserAuthorization,
    UserSpaceBrowserAuthRequest,
    UserSpaceBrowserAuthResponse,
    UserSpaceBrowserSurface,
    UserSpaceCapabilityTokenResponse,
    UserSpaceFileResponse,
    UserSpacePreviewLaunchRequest,
    UserSpacePreviewLaunchResponse,
    UserSpaceRuntimeActionResponse,
    UserSpaceRuntimeSessionResponse,
    UserSpaceRuntimeStatusResponse,
    UserSpaceWorkspaceTabStateResponse,
)
from ragtime.userspace.runtime_errors import RuntimeVersionConflictError
from ragtime.userspace.share_auth import set_share_auth_cookie, share_auth_token_from_request

logger = get_logger(__name__)

router = APIRouter(prefix="/indexes/userspace", tags=["User Space Runtime"])


_PROXY_METHODS = ["GET", "POST", "PUT", "PATCH", "DELETE", "HEAD", "OPTIONS"]
_COLLAB_CAPABILITY_COOKIE = "userspace_collab_capability"
_RUNTIME_PTY_CAPABILITY_COOKIE = "userspace_runtime_pty_capability"
_PROXY_TIMEOUT_FLOOR = 300.0  # seconds — minimum proxy read/write timeout
_PROXY_TIMEOUT_BUFFER = 20.0  # seconds — headroom above max tool timeout
_USERSPACE_SURFACE_HEADER = "X-Ragtime-Userspace-Surface"
_USERSPACE_PREVIEW_PROXY_HEADER = "X-Ragtime-Userspace-Preview-Proxy"
_DOCUMENT_PARSE_UPLOAD_CHUNK_BYTES = 1024 * 1024
_DOCUMENT_PARSE_MAX_BYTES = 25 * 1024 * 1024
_PRIMITIVE_UPLOAD_CHUNK_BYTES = 1024 * 1024
_PRIMITIVE_FILE_MAX_BYTES = USERSPACE_PRIMITIVE_UPLOAD_DEFAULT_MAX_BYTES
_PRIMITIVE_OBJECT_MAX_BYTES = USERSPACE_PRIMITIVE_UPLOAD_DEFAULT_MAX_BYTES
_PRIMITIVE_PROGRESS_STATES_MAX = 1000
_PRIMITIVE_PROGRESS_STATES: dict[tuple[str, str], dict[str, Any]] = {}
_PRIMITIVE_JOBS_MAX = 1000
_PRIMITIVE_JOBS: dict[tuple[str, str], dict[str, Any]] = {}

# Namespace prefix applied to user-app cookies as they cross the preview proxy
# boundary. The runtime worker rewrites every upstream ``Set-Cookie`` name to
# ``{_USER_APP_COOKIE_PREFIX}{base64url(original_name)}`` before it reaches the
# browser, and only cookies carrying this prefix are decoded back to the app on
# the way upstream. This makes platform cookies (preview session, capability,
# share-auth) structurally unforwardable to untrusted app code AND prevents a
# malicious app from shadowing a platform cookie by name — without maintaining a
# blocklist of platform cookie names (which cannot be enumerated because
# share-auth cookie names are derived from the share token at runtime).
#
# This is a COPY of the same constant in ``runtime/worker/api.py``. The ragtime
# app and runtime worker containers cannot cross-import, so the two definitions
# must be kept byte-for-byte in sync. Changing the prefix on only one side
# silently breaks user-app session persistence in previews.
_USER_APP_COOKIE_PREFIX = "__ragtime_app_cookie_"


def _decode_user_app_cookie_name(cookie_name: str) -> str | None:
    normalized = cookie_name.strip()
    if not normalized.startswith(_USER_APP_COOKIE_PREFIX):
        return None
    encoded = normalized[len(_USER_APP_COOKIE_PREFIX) :]
    if not encoded:
        return None
    try:
        padding = "=" * (-len(encoded) % 4)
        decoded = base64.urlsafe_b64decode((encoded + padding).encode("ascii")).decode("utf-8")
    except Exception:
        return None
    return decoded if decoded and "=" not in decoded and ";" not in decoded else None


def _sanitize_user_app_cookie_header(raw_cookie: str | None) -> str | None:
    if not raw_cookie:
        return None
    parts: list[str] = []
    for item in raw_cookie.split(";"):
        part = item.strip()
        if not part or "=" not in part:
            continue
        name, value = part.split("=", 1)
        decoded_name = _decode_user_app_cookie_name(name)
        if decoded_name:
            parts.append(f"{decoded_name}={value}")
    return "; ".join(parts) if parts else None


def _is_user_app_set_cookie(raw_set_cookie: str | None) -> bool:
    if not raw_set_cookie or "=" not in raw_set_cookie:
        return False
    name, _ = raw_set_cookie.split("=", 1)
    return _decode_user_app_cookie_name(name) is not None


class _CancellationSafeStreamingResponse(StreamingResponse):
    """Suppress shutdown/disconnect cancellation noise for long-lived streams."""

    async def __call__(self, scope: Any, receive: Any, send: Any) -> None:
        try:
            await super().__call__(scope, receive, send)
        except BaseException as exc:
            if not _is_cancellation_only(exc):
                raise
            logger.debug("Userspace streaming response cancelled during disconnect or shutdown")


def _is_cancellation_only(exc: BaseException) -> bool:
    if isinstance(exc, asyncio.CancelledError):
        return True
    if isinstance(exc, BaseExceptionGroup):
        return all(_is_cancellation_only(inner) for inner in exc.exceptions)
    return False


def _runtime_service() -> Any:
    # Lazy import avoids import cycle: runtime_service -> service -> preview_host -> runtime_routes.
    from ragtime.userspace.runtime_service import userspace_runtime_service

    return userspace_runtime_service


def _userspace_service() -> Any:
    # Lazy import avoids import cycle: service -> preview_host -> runtime_routes.
    from ragtime.userspace.service import userspace_service

    return userspace_service


def _root_share_target_url(base_url: str, share_path: str, target_path: str) -> str:
    normalized_base = (base_url or "").rstrip("/")
    normalized_share_path = "/" + share_path.lstrip("/")
    normalized_target_path = "/" + (target_path or "/").lstrip("/")
    if normalized_target_path == "/":
        return f"{normalized_base}{normalized_share_path}"
    return f"{normalized_base}{normalized_share_path.rstrip('/')}{normalized_target_path}"


def _workspace_preview_entry_url(
    base_url: str,
    workspace_id: str,
    target_path: str,
    parent_origin: str | None,
) -> str:
    normalized_base = (base_url or "").rstrip("/")
    entry_path = f"/indexes/userspace/runtime/workspaces/{quote(workspace_id, safe='')}/preview-entry"
    query_params = {"path": target_path or "/"}
    normalized_parent_origin = (parent_origin or "").strip()
    if normalized_parent_origin:
        query_params["parent_origin"] = normalized_parent_origin
    query = urlencode(query_params)
    return f"{normalized_base}{entry_path}?{query}"


def _root_share_launch_response(
    *,
    workspace_id: str,
    base_url: str,
    share_path: str,
    target_path: str,
) -> UserSpacePreviewLaunchResponse:
    preview_origin = (base_url or "").rstrip("/")
    return UserSpacePreviewLaunchResponse(
        workspace_id=workspace_id,
        preview_origin=preview_origin,
        preview_url=_root_share_target_url(base_url, share_path, target_path),
        expires_at=datetime.now(timezone.utc),
    )


def _workspace_preview_entry_launch_response(
    *,
    workspace_id: str,
    base_url: str,
    target_path: str,
    parent_origin: str | None,
    preview_origin: str,
    expires_at: datetime,
    preview_warning: Any | None = None,
) -> UserSpacePreviewLaunchResponse:
    return UserSpacePreviewLaunchResponse(
        workspace_id=workspace_id,
        preview_origin=preview_origin,
        preview_url=_workspace_preview_entry_url(
            base_url,
            workspace_id,
            target_path,
            parent_origin,
        ),
        expires_at=expires_at,
        preview_warning=preview_warning,
    )


def _preview_host_unreachable_response(
    *,
    workspace_id: str,
    preview_origin: str,
    warning: Any,
) -> HTMLResponse:
    """Return an actionable 503 when the preview subdomain is not routed to Ragtime.

    Mirrors the ``preview_host_unreachable`` warning condition but returns a
    local diagnostic page instead of issuing a 307 into an upstream that will 502 at the
    reverse proxy.
    """

    body = render_preview_host_unreachable_page_html(
        workspace_id=workspace_id,
        preview_origin=preview_origin,
        warning=warning,
    )
    return HTMLResponse(
        content=body,
        status_code=503,
        headers={
            "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
            "Pragma": "no-cache",
            "Referrer-Policy": "no-referrer",
            "X-Ragtime-Preview-Issue": "preview_host_unreachable",
        },
    )


def _collab_cookie_path(workspace_id: str) -> str:
    return f"/indexes/userspace/collab/workspaces/{workspace_id}/files"


def _runtime_pty_cookie_path(workspace_id: str) -> str:
    return f"/indexes/userspace/runtime/workspaces/{workspace_id}/pty"


_BROWSER_SURFACE_COOKIE_CONFIG: dict[str, dict[str, Any]] = {
    "collab": {
        "cookie_name": _COLLAB_CAPABILITY_COOKIE,
        "capabilities": ["userspace.collab_connect"],
        "path_builder": _collab_cookie_path,
    },
    "runtime_pty": {
        "cookie_name": _RUNTIME_PTY_CAPABILITY_COOKIE,
        "capabilities": ["userspace.runtime_pty"],
        "path_builder": _runtime_pty_cookie_path,
    },
}

_PREVIEW_SESSION_CAPABILITY = "userspace.preview_session"


def _default_browser_surfaces() -> list[UserSpaceBrowserSurface]:
    return ["collab", "runtime_pty"]


def _auth_user_payload(user: Any, *, user_fingerprint: str | None = None) -> dict[str, Any] | None:
    if user is None:
        return None
    username, display_name = normalize_user_identity(
        str(getattr(user, "username", "") or ""),
        str(getattr(user, "displayName", "") or getattr(user, "display_name", "") or ""),
    )
    display_label = display_name or username or None
    return {
        "display_name": display_label,
        "username": username or None,
        "email": getattr(user, "email", None),
        "auth_provider": str(getattr(getattr(user, "authProvider", None), "value", getattr(user, "authProvider", None)) or "local"),
        "user_fingerprint": user_fingerprint,
    }


def _normalize_browser_surfaces(
    raw_surfaces: Sequence[UserSpaceBrowserSurface] | None,
) -> list[UserSpaceBrowserSurface]:
    surfaces = raw_surfaces or _default_browser_surfaces()
    ordered: list[UserSpaceBrowserSurface] = []
    seen: set[UserSpaceBrowserSurface] = set()
    for surface in surfaces:
        if surface in _BROWSER_SURFACE_COOKIE_CONFIG and surface not in seen:
            ordered.append(surface)
            seen.add(surface)
    if raw_surfaces is None:
        return ordered or _default_browser_surfaces()
    return ordered


async def _userspace_auth_methods() -> list[UserSpaceAuthMethod]:
    binding_surfaces = _default_browser_surfaces()
    return [
        UserSpaceAuthMethod(
            **status,
            binding_surfaces=(binding_surfaces if bool(status.get("configured")) and bool(status.get("available")) else []),
        )
        for status in await build_auth_method_statuses()
    ]


async def _validate_auth_binding_method(auth_method_key: str | None) -> str | None:
    normalized_key = normalize_auth_method_key(auth_method_key)
    if not normalized_key:
        return None
    methods = await _userspace_auth_methods()
    methods_by_key = {method.key: method for method in methods}
    method = methods_by_key.get(normalized_key)
    if method is None:
        raise HTTPException(status_code=400, detail=f"Unknown auth method: {normalized_key}")
    if not method.configured:
        raise HTTPException(status_code=400, detail=f"Auth method is not configured: {normalized_key}")
    if not method.available:
        raise HTTPException(status_code=400, detail=f"Auth method is not available: {normalized_key}")
    return normalized_key


async def _primitive_session_payload(
    workspace_id: str,
    user_id: str | None,
    *,
    mode: str,
    share_access_mode: Any = None,
    user: Any = None,
    same_origin_auth_endpoints: bool = False,
) -> dict[str, Any]:
    capabilities = await _primitive_capabilities(workspace_id, user_id, preview_mode=mode)
    fingerprint = None
    if user_id:
        fingerprint = (
            build_workspace_user_fingerprint(
                user_id=user_id,
                workspace_id=workspace_id,
                workspace_fingerprint_namespace=await _userspace_service().get_workspace_audit_fingerprint_namespace(workspace_id),
                user_identity_subject=build_user_fingerprint_subject(
                    user_id=user_id,
                    username=getattr(user, "username", None),
                    auth_provider=getattr(user, "authProvider", None),
                    ldap_dn=getattr(user, "ldapDn", None),
                    source_provider=getattr(user, "sourceProvider", None),
                    source_id=getattr(user, "sourceId", None),
                ),
            )
            or None
        )
    auth_methods = await _userspace_auth_methods()
    browser_auth_endpoint = "/__ragtime/browser-auth"
    browser_logout_endpoint = "/__ragtime/browser-auth/logout"
    if mode == "workspace" and not same_origin_auth_endpoints:
        browser_auth_endpoint = f"/indexes/userspace/runtime/workspaces/{workspace_id}/browser-auth"
        browser_logout_endpoint = f"/indexes/userspace/runtime/workspaces/{workspace_id}/browser-auth/logout"
    return {
        "workspace_id": workspace_id,
        "mode": mode,
        "user_id": user_id,
        "user_fingerprint": fingerprint,
        "user_fingerprint_scope": USER_FINGERPRINT_SCOPE_WORKSPACE if fingerprint else None,
        "share_access_mode": share_access_mode,
        "auth": {
            "authenticated": bool(user_id),
            "user": _auth_user_payload(user, user_fingerprint=fingerprint) if user_id else None,
            "methods": [method.model_dump() for method in auth_methods],
            "binding_strategy": "browser_capability_token",
            "interactive_auth_endpoint": "/__ragtime/browser-auth/start",
            "login_endpoint": "/auth/login",
            "current_user_endpoint": "/auth/me",
            "browser_auth_endpoint": browser_auth_endpoint,
            "browser_logout_endpoint": browser_logout_endpoint,
        },
        "capabilities": capabilities,
    }


def _get_max_proxy_timeout() -> float:
    """Return the proxy read/write timeout derived from max tool timeout_max_seconds."""
    try:
        from ragtime.core.app_settings import SettingsCache

        cached_configs = SettingsCache.get_instance()._tool_configs
        if cached_configs:
            max_t = max(
                (cfg.get("timeout_max_seconds", 300) or 300 for cfg in cached_configs),
                default=300,
            )
            return max(float(max_t), _PROXY_TIMEOUT_FLOOR) + _PROXY_TIMEOUT_BUFFER
    except Exception:
        pass
    return _PROXY_TIMEOUT_FLOOR + _PROXY_TIMEOUT_BUFFER


async def _safe_close_websocket(websocket: WebSocket, code: int) -> None:
    with contextlib.suppress(Exception):
        await websocket.close(code=code)


def _safe_upload_filename(file: StarletteUploadFile) -> str:
    raw_name = (file.filename or "upload").strip().replace("\\", "/")
    filename = Path(raw_name).name or "upload"
    if filename in {".", ".."}:
        filename = "upload"
    return filename


async def _get_primitive_upload_max_bytes() -> int:
    try:
        app_settings = await get_app_settings()
        return clamp_userspace_primitive_upload_max_bytes(app_settings.get("userspace_primitive_upload_max_bytes"))
    except Exception:
        return _PRIMITIVE_FILE_MAX_BYTES


async def _get_primitive_archive_max_entries() -> int:
    try:
        app_settings = await get_app_settings()
        return clamp_userspace_primitive_archive_max_entries(app_settings.get("userspace_primitive_archive_max_entries"))
    except Exception:
        return USERSPACE_PRIMITIVE_ARCHIVE_DEFAULT_MAX_ENTRIES


def _format_limit_bytes(value: int) -> str:
    if value % (1024 * 1024 * 1024) == 0:
        return f"{value // (1024 * 1024 * 1024)} GB ({value:,} bytes)"
    if value % (1024 * 1024) == 0:
        return f"{value // (1024 * 1024)} MB ({value:,} bytes)"
    return f"{value:,} bytes"


def _limit_exceeded_detail(label: str, max_bytes: int, limit_name: str) -> str:
    return f"{label} exceeds the {limit_name} of {_format_limit_bytes(max_bytes)}."


def _archive_entry_limit_detail(max_entries: int) -> str:
    return f"Archive contains more files than the configured User Space primitive archive file limit of {max_entries:,}."


def _archive_extracted_bytes_limit_detail(max_bytes: int) -> str:
    return f"Extracted archive content exceeds the configured User Space primitive upload size limit of {_format_limit_bytes(max_bytes)}."


def _archive_member_bytes_limit_detail(max_bytes: int) -> str:
    return f"Archive member exceeds the configured User Space primitive upload size limit of {_format_limit_bytes(max_bytes)}."


async def _read_upload_bytes(
    file: StarletteUploadFile,
    *,
    max_bytes: int,
    label: str,
    limit_name: str,
) -> tuple[str, str | None, int, bytes]:
    filename = _safe_upload_filename(file)

    content_chunks: list[bytes] = []
    total_bytes = 0
    try:
        while True:
            chunk = await file.read(_PRIMITIVE_UPLOAD_CHUNK_BYTES)
            if not chunk:
                break
            total_bytes += len(chunk)
            if total_bytes > max_bytes:
                raise HTTPException(
                    status_code=413,
                    detail=_limit_exceeded_detail(label, max_bytes, limit_name),
                )
            content_chunks.append(chunk)
    finally:
        with contextlib.suppress(Exception):
            await file.close()

    return filename, file.content_type or None, total_bytes, b"".join(content_chunks)


async def _read_request_bytes(
    request: Request,
    *,
    max_bytes: int,
    label: str,
    limit_name: str,
) -> tuple[str | None, int, bytes]:
    content_type = request.headers.get("content-type") or None
    if content_type and content_type.lower().startswith("multipart/form-data"):
        form = await request.form()
        form_file = form.get("file")
        if not isinstance(form_file, StarletteUploadFile):
            raise HTTPException(status_code=400, detail="Multipart request must include file field")
        _, upload_content_type, total_bytes, content = await _read_upload_bytes(
            form_file,
            max_bytes=max_bytes,
            label=label,
            limit_name=limit_name,
        )
        return upload_content_type, total_bytes, content

    content_chunks: list[bytes] = []
    total_bytes = 0
    async for chunk in request.stream():
        if not chunk:
            continue
        total_bytes += len(chunk)
        if total_bytes > max_bytes:
            raise HTTPException(status_code=413, detail=_limit_exceeded_detail(label, max_bytes, limit_name))
        content_chunks.append(chunk)
    return content_type, total_bytes, b"".join(content_chunks)


async def _parse_uploaded_document_upload(file: UploadFile) -> dict[str, Any]:
    filename, content_type, total_bytes, content = await _read_upload_bytes(
        file,
        max_bytes=_DOCUMENT_PARSE_MAX_BYTES,
        label="Uploaded file",
        limit_name="platform document parsing upload size limit",
    )
    extracted_text = await extract_text_from_file_async(
        Path(filename),
        content=content,
    )
    return {
        "filename": filename,
        "content_type": content_type,
        "size_bytes": total_bytes,
        "text": extracted_text,
    }


def _spreadsheet_cell_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value)


def _spreadsheet_text_from_sheets(sheets: list[dict[str, Any]]) -> str:
    parts: list[str] = []
    for sheet in sheets:
        name = str(sheet.get("name") or "Sheet")
        parts.append(f"## Sheet: {name}")
        rows_value = sheet.get("rows")
        rows = rows_value if isinstance(rows_value, list) else []
        for row in rows:
            if not isinstance(row, list):
                continue
            row_text = " | ".join(str(value) for value in row if value not in (None, ""))
            if row_text:
                parts.append(row_text)
    return "\n".join(parts)


def _extract_xlsx_sheets(content: bytes) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise HTTPException(status_code=415, detail="XLSX parsing requires openpyxl") from exc

    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Invalid XLSX workbook") from exc
    try:
        sheets: list[dict[str, Any]] = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            rows: list[list[Any]] = []
            for row in sheet.iter_rows(values_only=True):
                values = [_spreadsheet_cell_value(value) for value in row]
                while values and values[-1] is None:
                    values.pop()
                if values:
                    rows.append(values)
            sheets.append({"name": sheet_name, "rows": rows})
        return sheets
    finally:
        workbook.close()


def _extract_xls_sheets(content: bytes) -> list[dict[str, Any]]:
    try:
        import xlrd
    except ImportError as exc:
        raise HTTPException(status_code=415, detail="XLS parsing requires xlrd") from exc

    try:
        workbook = xlrd.open_workbook(file_contents=content)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Invalid XLS workbook") from exc

    sheets: list[dict[str, Any]] = []
    for sheet in workbook.sheets():
        rows: list[list[Any]] = []
        for row_index in range(sheet.nrows):
            values = [_spreadsheet_cell_value(sheet.cell_value(row_index, column_index)) for column_index in range(sheet.ncols)]
            while values and values[-1] in (None, ""):
                values.pop()
            if values:
                rows.append(values)
        sheets.append({"name": sheet.name, "rows": rows})
    return sheets


def _extract_csv_sheet(content: bytes) -> list[dict[str, Any]]:
    text = content.decode("utf-8-sig", errors="replace")
    rows = [row for row in csv.reader(io.StringIO(text)) if any(cell for cell in row)]
    return [{"name": "CSV", "rows": rows}]


async def _parse_uploaded_spreadsheet_upload(file: UploadFile) -> dict[str, Any]:
    filename, content_type, total_bytes, content = await _read_upload_bytes(
        file,
        max_bytes=_DOCUMENT_PARSE_MAX_BYTES,
        label="Uploaded spreadsheet",
        limit_name="platform spreadsheet parsing upload size limit",
    )
    suffix = Path(filename).suffix.lower()
    if suffix == ".xlsx":
        sheets = await asyncio.to_thread(_extract_xlsx_sheets, content)
    elif suffix == ".xls":
        sheets = await asyncio.to_thread(_extract_xls_sheets, content)
    elif suffix == ".csv":
        sheets = await asyncio.to_thread(_extract_csv_sheet, content)
    else:
        raise HTTPException(status_code=415, detail="Supported spreadsheet formats are XLSX, XLS, and CSV")

    return {
        "filename": filename,
        "content_type": content_type,
        "size_bytes": total_bytes,
        "sheets": sheets,
        "text": _spreadsheet_text_from_sheets(sheets),
    }


def _infer_table_column_type(values: list[Any]) -> str:
    typed_values = [value for value in values if value not in (None, "")]
    if not typed_values:
        return "empty"
    if all(isinstance(value, bool) for value in typed_values):
        return "boolean"
    if all(isinstance(value, (int, float)) and not isinstance(value, bool) for value in typed_values):
        return "number"
    if all(isinstance(value, str) for value in typed_values):
        return "string"
    return "mixed"


def _normalize_table_headers(first_row: list[Any], column_count: int) -> tuple[list[str], bool]:
    raw_headers = [str(value or "").strip() for value in first_row[:column_count]]
    seen: set[str] = set()
    headers: list[str] = []
    has_explicit_headers = bool(raw_headers) and all(raw_headers) and len(set(raw_headers)) == len(raw_headers)
    for index in range(column_count):
        candidate = raw_headers[index] if index < len(raw_headers) and raw_headers[index] else f"column_{index + 1}"
        header = re.sub(r"\s+", "_", candidate.strip().lower()) or f"column_{index + 1}"
        header = re.sub(r"[^a-z0-9_]+", "_", header).strip("_") or f"column_{index + 1}"
        if header in seen:
            suffix = 2
            base = header
            while f"{base}_{suffix}" in seen:
                suffix += 1
            header = f"{base}_{suffix}"
        seen.add(header)
        headers.append(header)
    return headers, has_explicit_headers


def _normalize_spreadsheet_tables(sheets: list[dict[str, Any]]) -> list[dict[str, Any]]:
    normalized_sheets: list[dict[str, Any]] = []
    for sheet in sheets:
        rows_value = sheet.get("rows")
        rows = [row for row in rows_value if isinstance(row, list)] if isinstance(rows_value, list) else []
        column_count = max((len(row) for row in rows), default=0)
        headers, has_explicit_headers = _normalize_table_headers(rows[0] if rows else [], column_count)
        data_rows = rows[1:] if has_explicit_headers else rows
        records: list[dict[str, Any]] = []
        columns = [{"key": header, "label": header.replace("_", " ").title()} for header in headers]
        for row in data_rows:
            record: dict[str, Any] = {}
            for index, header in enumerate(headers):
                record[header] = row[index] if index < len(row) else None
            if any(value not in (None, "") for value in record.values()):
                records.append(record)
        for column in columns:
            key = str(column["key"])
            column["type"] = _infer_table_column_type([record.get(key) for record in records])
        normalized_sheets.append(
            {
                "name": str(sheet.get("name") or "Sheet"),
                "has_header_row": has_explicit_headers,
                "columns": columns,
                "records": records,
                "row_count": len(records),
            }
        )
    return normalized_sheets


async def _normalize_uploaded_table_upload(file: UploadFile) -> dict[str, Any]:
    parsed = await _parse_uploaded_spreadsheet_upload(file)
    sheets = parsed.get("sheets") if isinstance(parsed, dict) else []
    parsed["tables"] = _normalize_spreadsheet_tables(sheets if isinstance(sheets, list) else [])
    return parsed


def _render_document_preview_html(filename: str, text: str) -> str:
    title = _html.escape(filename or "Document preview")
    escaped_text = _html.escape(text or "")
    body = escaped_text.replace("\n", "<br>\n")
    return (
        '<!doctype html><html><head><meta charset="utf-8">'
        f"<title>{title}</title>"
        "<style>body{font-family:system-ui,-apple-system,Segoe UI,sans-serif;margin:24px;line-height:1.5;color:#222}"
        "pre{white-space:pre-wrap;font:inherit}</style></head><body>"
        f"<h1>{title}</h1><pre>{body}</pre></body></html>"
    )


async def _render_uploaded_document_preview_upload(file: UploadFile) -> dict[str, Any]:
    parsed = await _parse_uploaded_document_upload(file)
    filename = str(parsed.get("filename") or "upload")
    text = str(parsed.get("text") or "")
    return {
        **parsed,
        "format": "html",
        "html": _render_document_preview_html(filename, text),
    }


def _normalize_archive_member_path(member_name: str) -> str:
    raw_path = str(member_name or "").replace("\\", "/").strip("/")
    parts = PurePosixPath(raw_path).parts
    if not raw_path or any(part in {"", ".", ".."} for part in parts):
        raise HTTPException(status_code=400, detail="Archive contains an unsafe path")
    return "/".join(parts)


def _extract_archive_entries(content: bytes, *, max_entries: int, max_extracted_bytes: int) -> list[tuple[str, bytes]]:
    entries: list[tuple[str, bytes]] = []
    extracted_bytes = 0

    def _append_entry(name: str, payload: bytes) -> None:
        nonlocal extracted_bytes
        if len(entries) >= max_entries:
            raise HTTPException(status_code=413, detail=_archive_entry_limit_detail(max_entries))
        extracted_bytes += len(payload)
        if extracted_bytes > max_extracted_bytes:
            raise HTTPException(status_code=413, detail=_archive_extracted_bytes_limit_detail(max_extracted_bytes))
        entries.append((_normalize_archive_member_path(name), payload))

    buffer = io.BytesIO(content)
    if zipfile.is_zipfile(buffer):
        with zipfile.ZipFile(buffer) as zip_archive:
            for info in zip_archive.infolist():
                if info.is_dir():
                    continue
                if info.file_size > max_extracted_bytes:
                    raise HTTPException(status_code=413, detail=_archive_member_bytes_limit_detail(max_extracted_bytes))
                _append_entry(info.filename, zip_archive.read(info))
        return entries

    buffer.seek(0)
    try:
        with tarfile.open(fileobj=buffer, mode="r:*") as tar_archive:
            for member in tar_archive.getmembers():
                if not member.isfile():
                    continue
                if member.size > max_extracted_bytes:
                    raise HTTPException(status_code=413, detail=_archive_member_bytes_limit_detail(max_extracted_bytes))
                member_file = tar_archive.extractfile(member)
                if member_file is None:
                    continue
                _append_entry(member.name, member_file.read())
        return entries
    except tarfile.TarError as exc:
        raise HTTPException(status_code=415, detail="Supported archive formats are ZIP, TAR, TAR.GZ, and TGZ") from exc


def _join_primitive_relative_path(base_path: str, member_path: str) -> str:
    normalized_base = str(base_path or "").replace("\\", "/").strip("/")
    if not normalized_base:
        return member_path
    base_parts = PurePosixPath(normalized_base).parts
    if any(part in {"", ".", ".."} for part in base_parts):
        raise HTTPException(status_code=400, detail="Destination path must be relative")
    return "/".join([*base_parts, member_path])


async def _primitive_archive_extract(
    workspace_id: str,
    request: Request,
    user_id: str,
    *,
    target: str = "files",
    destination_path: str = "",
    bucket_name: str | None = None,
) -> dict[str, Any]:
    userspace_service = _userspace_service()
    await userspace_service.enforce_workspace_role(workspace_id, user_id, "editor")
    max_upload_bytes = await _get_primitive_upload_max_bytes()
    max_entries = await _get_primitive_archive_max_entries()
    content_type, total_bytes, content = await _read_request_bytes(
        request,
        max_bytes=max_upload_bytes,
        label="Uploaded archive",
        limit_name="configured User Space primitive upload size limit",
    )
    entries = await asyncio.to_thread(
        _extract_archive_entries,
        content,
        max_entries=max_entries,
        max_extracted_bytes=max_upload_bytes,
    )
    normalized_target = str(target or "files").strip().lower()
    written: list[dict[str, Any]] = []
    if normalized_target == "files":
        for member_path, payload in entries:
            output_path = _join_primitive_relative_path(destination_path, member_path)
            result = await _primitive_workspace_file_write_content(
                workspace_id,
                output_path,
                mimetypes.guess_type(output_path)[0] or "application/octet-stream",
                len(payload),
                payload,
                user_id,
            )
            written.append(result)
    elif normalized_target == "objects":
        bucket = str(bucket_name or "").strip()
        if not bucket:
            raise HTTPException(status_code=400, detail="bucket is required when target=objects")
        for member_path, payload in entries:
            object_path = _join_primitive_relative_path(destination_path, member_path)
            result = await _primitive_object_write_content(
                workspace_id,
                bucket,
                object_path,
                mimetypes.guess_type(object_path)[0] or "application/octet-stream",
                len(payload),
                payload,
                user_id,
            )
            written.append(result)
    else:
        raise HTTPException(status_code=400, detail="target must be 'files' or 'objects'")
    return {
        "target": normalized_target,
        "content_type": content_type,
        "archive_size_bytes": total_bytes,
        "extracted_count": len(written),
        "max_entries": max_entries,
        "items": written,
    }


def _normalize_primitive_object_key(object_path: str) -> str:
    raw_path = str(object_path or "").replace("\\", "/").strip("/")
    parts = PurePosixPath(raw_path).parts
    if not raw_path or not parts or any(part in {"", ".", ".."} for part in parts):
        raise HTTPException(status_code=400, detail="Object key must be a relative path")
    return "/".join(parts)


def _normalize_progress_task_id(task_id: str) -> str:
    normalized = str(task_id or "").strip()
    if not re.fullmatch(r"[A-Za-z0-9][A-Za-z0-9._:-]{0,127}", normalized):
        raise HTTPException(status_code=400, detail="Progress task id is invalid")
    return normalized


async def _primitive_workspace_file_response(workspace_id: str, file_path: str, user_id: str) -> Response:
    userspace_service = _userspace_service()
    await userspace_service.enforce_workspace_role(workspace_id, user_id, "viewer")
    normalized_path = _runtime_service()._normalize_file_path(file_path)
    normalized_path = await userspace_service.ensure_workspace_path_not_in_disabled_mount(workspace_id, normalized_path)
    if await userspace_service._is_workspace_mount_owned_path(workspace_id, normalized_path):
        raise HTTPException(status_code=409, detail="Binary file primitive does not read mounted paths")
    root = Path(userspace_service._workspace_files_dir(workspace_id)).resolve()
    target = (root / normalized_path).resolve()
    try:
        target.relative_to(root)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="File path must stay inside the workspace") from exc
    if not target.exists() or not target.is_file():
        raise HTTPException(status_code=404, detail="File not found")
    content = await asyncio.to_thread(target.read_bytes)
    media_type = mimetypes.guess_type(target.name)[0] or "application/octet-stream"
    return Response(
        content=content,
        media_type=media_type,
        headers={"X-Ragtime-File-Path": normalized_path},
    )


async def _primitive_workspace_file_write(
    workspace_id: str,
    file_path: str,
    file: UploadFile,
    user_id: str,
) -> dict[str, Any]:
    max_upload_bytes = await _get_primitive_upload_max_bytes()
    _, content_type, total_bytes, content = await _read_upload_bytes(
        file,
        max_bytes=max_upload_bytes,
        label="Uploaded workspace file",
        limit_name="configured User Space primitive upload size limit",
    )
    return await _primitive_workspace_file_write_content(
        workspace_id,
        file_path,
        content_type,
        total_bytes,
        content,
        user_id,
    )


async def _primitive_workspace_file_write_request(
    workspace_id: str,
    file_path: str,
    request: Request,
    user_id: str,
) -> dict[str, Any]:
    max_upload_bytes = await _get_primitive_upload_max_bytes()
    content_type, total_bytes, content = await _read_request_bytes(
        request,
        max_bytes=max_upload_bytes,
        label="Uploaded workspace file",
        limit_name="configured User Space primitive upload size limit",
    )
    return await _primitive_workspace_file_write_content(
        workspace_id,
        file_path,
        content_type,
        total_bytes,
        content,
        user_id,
    )


async def _primitive_workspace_file_write_content(
    workspace_id: str,
    file_path: str,
    content_type: str | None,
    total_bytes: int,
    content: bytes,
    user_id: str,
) -> dict[str, Any]:
    userspace_service = _userspace_service()
    await userspace_service.enforce_workspace_role(workspace_id, user_id, "editor")
    normalized_path = _runtime_service()._normalize_file_path(file_path)
    normalized_path = await userspace_service.ensure_workspace_path_not_in_disabled_mount(workspace_id, normalized_path)
    if await userspace_service._is_workspace_mount_owned_path(workspace_id, normalized_path):
        raise HTTPException(status_code=409, detail="Binary file primitive does not write mounted paths")
    root = Path(userspace_service._workspace_files_dir(workspace_id)).resolve()
    target = (root / normalized_path).resolve()
    try:
        target.relative_to(root)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="File path must stay inside the workspace") from exc
    target.parent.mkdir(parents=True, exist_ok=True)
    await asyncio.to_thread(target.write_bytes, content)
    await userspace_service.clear_workspace_changed_file_acknowledgements_for_paths_for_all_users(workspace_id, [normalized_path])
    await userspace_service.touch_workspace(workspace_id)
    return {
        "path": normalized_path,
        "size_bytes": total_bytes,
        "content_type": content_type,
    }


def _object_storage_target(workspace_id: str, bucket_name: str, object_path: str) -> tuple[str, str, Path]:
    userspace_service = _userspace_service()
    bucket = userspace_service._normalize_object_storage_bucket_name(bucket_name)
    key = _normalize_primitive_object_key(object_path)
    payload = userspace_service._ensure_object_storage_config(workspace_id)
    buckets_value = payload.get("buckets") if isinstance(payload, dict) else []
    buckets = buckets_value if isinstance(buckets_value, list) else []
    if not any(isinstance(item, dict) and str(item.get("name") or "") == bucket for item in buckets if item):
        raise HTTPException(status_code=404, detail="Object storage bucket not found")
    root = Path(userspace_service._workspace_object_storage_buckets_dir(workspace_id)).resolve()
    target = (root / bucket / key).resolve()
    try:
        target.relative_to(root / bucket)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Object key must stay inside the bucket") from exc
    return bucket, key, target


def _primitive_object_url(bucket: str, key: str) -> str:
    return f"/__ragtime/objects/{quote(bucket, safe='')}/{quote(key, safe='/')}"


def _runtime_primitive_file_url(workspace_id: str, file_path: str) -> str:
    return f"/indexes/userspace/runtime/workspaces/{quote(workspace_id, safe='')}/primitives/files/{quote(file_path, safe='/')}"


def _runtime_primitive_object_url(workspace_id: str, bucket: str, key: str) -> str:
    return f"/indexes/userspace/runtime/workspaces/{quote(workspace_id, safe='')}/primitives/objects/{quote(bucket, safe='')}/{quote(key, safe='/')}"


async def _primitive_object_response(workspace_id: str, bucket_name: str, object_path: str, user_id: str) -> Response:
    await _userspace_service().enforce_workspace_role(workspace_id, user_id, "viewer")
    bucket, key, target = _object_storage_target(workspace_id, bucket_name, object_path)
    if not target.exists() or not target.is_file():
        raise HTTPException(status_code=404, detail="Object not found")
    content = await asyncio.to_thread(target.read_bytes)
    media_type = mimetypes.guess_type(target.name)[0] or "application/octet-stream"
    return Response(
        content=content,
        media_type=media_type,
        headers={
            "X-Ragtime-Object-Bucket": bucket,
            "X-Ragtime-Object-Key": key,
        },
    )


async def _primitive_object_write(
    workspace_id: str,
    bucket_name: str,
    object_path: str,
    file: UploadFile,
    user_id: str,
) -> dict[str, Any]:
    max_upload_bytes = await _get_primitive_upload_max_bytes()
    _, content_type, total_bytes, content = await _read_upload_bytes(
        file,
        max_bytes=max_upload_bytes,
        label="Uploaded object",
        limit_name="configured User Space primitive upload size limit",
    )
    return await _primitive_object_write_content(
        workspace_id,
        bucket_name,
        object_path,
        content_type,
        total_bytes,
        content,
        user_id,
    )


async def _primitive_object_write_request(
    workspace_id: str,
    bucket_name: str,
    object_path: str,
    request: Request,
    user_id: str,
) -> dict[str, Any]:
    max_upload_bytes = await _get_primitive_upload_max_bytes()
    content_type, total_bytes, content = await _read_request_bytes(
        request,
        max_bytes=max_upload_bytes,
        label="Uploaded object",
        limit_name="configured User Space primitive upload size limit",
    )
    return await _primitive_object_write_content(
        workspace_id,
        bucket_name,
        object_path,
        content_type,
        total_bytes,
        content,
        user_id,
    )


async def _primitive_object_write_content(
    workspace_id: str,
    bucket_name: str,
    object_path: str,
    content_type: str | None,
    total_bytes: int,
    content: bytes,
    user_id: str,
) -> dict[str, Any]:
    await _userspace_service().enforce_workspace_role(workspace_id, user_id, "editor")
    bucket, key, target = _object_storage_target(workspace_id, bucket_name, object_path)
    target.parent.mkdir(parents=True, exist_ok=True)
    await asyncio.to_thread(target.write_bytes, content)
    await _userspace_service().touch_workspace(workspace_id)
    return {
        "bucket": bucket,
        "key": key,
        "size_bytes": total_bytes,
        "content_type": content_type,
        "url": _primitive_object_url(bucket, key),
    }


async def _primitive_progress_get(workspace_id: str, task_id: str, user_id: str) -> dict[str, Any]:
    await _userspace_service().enforce_workspace_role(workspace_id, user_id, "viewer")
    normalized_task_id = _normalize_progress_task_id(task_id)
    state = _PRIMITIVE_PROGRESS_STATES.get((workspace_id, normalized_task_id))
    if state is None:
        raise HTTPException(status_code=404, detail="Progress task not found")
    return dict(state)


async def _primitive_progress_put(
    workspace_id: str,
    task_id: str,
    payload: dict[str, Any],
    user_id: str,
) -> dict[str, Any]:
    await _userspace_service().enforce_workspace_role(workspace_id, user_id, "editor")
    normalized_task_id = _normalize_progress_task_id(task_id)
    raw_progress = payload.get("progress") if isinstance(payload, dict) else None
    progress = None
    if raw_progress is not None:
        try:
            progress = max(0.0, min(1.0, float(raw_progress)))
        except (TypeError, ValueError) as exc:
            raise HTTPException(status_code=400, detail="Progress must be a number between 0 and 1") from exc
    if len(_PRIMITIVE_PROGRESS_STATES) >= _PRIMITIVE_PROGRESS_STATES_MAX:
        oldest_key = min(_PRIMITIVE_PROGRESS_STATES, key=lambda key: str(_PRIMITIVE_PROGRESS_STATES[key].get("updated_at") or ""))
        _PRIMITIVE_PROGRESS_STATES.pop(oldest_key, None)
    state = {
        "workspace_id": workspace_id,
        "task_id": normalized_task_id,
        "phase": str(payload.get("phase") or "running") if isinstance(payload, dict) else "running",
        "progress": progress,
        "message": str(payload.get("message") or "") if isinstance(payload, dict) else "",
        "data": payload.get("data") if isinstance(payload, dict) else None,
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    _PRIMITIVE_PROGRESS_STATES[(workspace_id, normalized_task_id)] = state
    return dict(state)


def _normalize_job_id(job_id: str) -> str:
    normalized = str(job_id or "").strip()
    if not re.fullmatch(r"[A-Za-z0-9][A-Za-z0-9._:-]{0,127}", normalized):
        raise HTTPException(status_code=400, detail="Job id is invalid")
    return normalized


async def _primitive_job_get(workspace_id: str, job_id: str, user_id: str) -> dict[str, Any]:
    await _userspace_service().enforce_workspace_role(workspace_id, user_id, "viewer")
    normalized_job_id = _normalize_job_id(job_id)
    job = _PRIMITIVE_JOBS.get((workspace_id, normalized_job_id))
    if job is None:
        raise HTTPException(status_code=404, detail="Job not found")
    return dict(job)


async def _primitive_job_put(workspace_id: str, job_id: str, payload: dict[str, Any], user_id: str) -> dict[str, Any]:
    await _userspace_service().enforce_workspace_role(workspace_id, user_id, "editor")
    normalized_job_id = _normalize_job_id(job_id)
    if len(_PRIMITIVE_JOBS) >= _PRIMITIVE_JOBS_MAX and (workspace_id, normalized_job_id) not in _PRIMITIVE_JOBS:
        oldest_key = min(_PRIMITIVE_JOBS, key=lambda key: str(_PRIMITIVE_JOBS[key].get("updated_at") or ""))
        _PRIMITIVE_JOBS.pop(oldest_key, None)
    existing = _PRIMITIVE_JOBS.get((workspace_id, normalized_job_id), {})
    now = datetime.now(timezone.utc).isoformat()
    job = {
        "workspace_id": workspace_id,
        "job_id": normalized_job_id,
        "kind": str(payload.get("kind") or existing.get("kind") or "generic") if isinstance(payload, dict) else "generic",
        "status": str(payload.get("status") or existing.get("status") or "running") if isinstance(payload, dict) else "running",
        "progress": payload.get("progress", existing.get("progress")) if isinstance(payload, dict) else existing.get("progress"),
        "message": str(payload.get("message") or "") if isinstance(payload, dict) else "",
        "data": payload.get("data") if isinstance(payload, dict) else None,
        "result": payload.get("result") if isinstance(payload, dict) else None,
        "error": payload.get("error") if isinstance(payload, dict) else None,
        "created_at": existing.get("created_at") or now,
        "updated_at": now,
    }
    _PRIMITIVE_JOBS[(workspace_id, normalized_job_id)] = job
    return dict(job)


async def _primitive_job_create(workspace_id: str, payload: dict[str, Any], user_id: str) -> dict[str, Any]:
    requested_id = str(payload.get("job_id") or "").strip() if isinstance(payload, dict) else ""
    job_id = requested_id or f"job-{uuid.uuid4().hex[:16]}"
    return await _primitive_job_put(workspace_id, job_id, payload if isinstance(payload, dict) else {}, user_id)


async def _primitive_upload_target(
    workspace_id: str,
    payload: dict[str, Any],
    user_id: str,
    *,
    preview_origin: bool,
) -> dict[str, Any]:
    await _userspace_service().enforce_workspace_role(workspace_id, user_id, "editor")
    target = str(payload.get("target") or payload.get("storage") or "files").strip().lower()
    max_upload_bytes = await _get_primitive_upload_max_bytes()
    if target == "files":
        raw_path = str(payload.get("path") or payload.get("file_path") or "").strip()
        if not raw_path:
            raise HTTPException(status_code=400, detail="path is required for file upload targets")
        normalized_path = _runtime_service()._normalize_file_path(raw_path)
        normalized_path = await _userspace_service().ensure_workspace_path_not_in_disabled_mount(workspace_id, normalized_path)
        if await _userspace_service()._is_workspace_mount_owned_path(workspace_id, normalized_path):
            raise HTTPException(status_code=409, detail="Upload target cannot point at mounted paths")
        url = f"/__ragtime/files/{quote(normalized_path, safe='/')}" if preview_origin else _runtime_primitive_file_url(workspace_id, normalized_path)
        return {
            "target": "files",
            "method": "PUT",
            "url": url,
            "path": normalized_path,
            "max_bytes": max_upload_bytes,
            "headers": {},
        }
    if target == "objects":
        bucket = str(payload.get("bucket") or payload.get("bucket_name") or "").strip()
        key = str(payload.get("key") or payload.get("object_path") or "").strip()
        if not bucket or not key:
            raise HTTPException(status_code=400, detail="bucket and key are required for object upload targets")
        normalized_bucket, normalized_key, _ = _object_storage_target(workspace_id, bucket, key)
        url = (
            _primitive_object_url(normalized_bucket, normalized_key)
            if preview_origin
            else _runtime_primitive_object_url(workspace_id, normalized_bucket, normalized_key)
        )
        return {
            "target": "objects",
            "method": "PUT",
            "url": url,
            "bucket": normalized_bucket,
            "key": normalized_key,
            "max_bytes": max_upload_bytes,
            "headers": {},
        }
    raise HTTPException(status_code=400, detail="target must be 'files' or 'objects'")


async def _primitive_capabilities(
    workspace_id: str,
    user_id: str | None,
    *,
    preview_mode: str = "workspace",
) -> dict[str, Any]:
    max_upload_bytes = await _get_primitive_upload_max_bytes()
    max_archive_entries = await _get_primitive_archive_max_entries()
    can_read_workspace = False
    can_write_workspace = False
    buckets: list[str] = []
    if user_id and preview_mode == "workspace":
        try:
            await _userspace_service().enforce_workspace_role(workspace_id, user_id, "viewer")
            can_read_workspace = True
        except HTTPException:
            can_read_workspace = False
        try:
            await _userspace_service().enforce_workspace_role(workspace_id, user_id, "editor")
            can_write_workspace = True
        except HTTPException:
            can_write_workspace = False
        if can_read_workspace:
            with contextlib.suppress(Exception):
                payload = _userspace_service()._ensure_object_storage_config(workspace_id)
                bucket_items = payload.get("buckets") if isinstance(payload, dict) else []
                if not isinstance(bucket_items, list):
                    bucket_items = []
                buckets = [str(item.get("name")) for item in bucket_items if isinstance(item, dict) and item.get("name")]
    return {
        "workspace_id": workspace_id,
        "mode": preview_mode,
        "can_parse": True,
        "can_render_preview": True,
        "can_normalize_tables": True,
        "can_read_files": can_read_workspace,
        "can_write_files": can_write_workspace,
        "can_read_objects": can_read_workspace,
        "can_write_objects": can_write_workspace,
        "can_extract_archives": can_write_workspace,
        "can_create_upload_targets": can_write_workspace,
        "can_read_progress": can_read_workspace,
        "can_write_progress": can_write_workspace,
        "can_read_jobs": can_read_workspace,
        "can_write_jobs": can_write_workspace,
        "max_upload_bytes": max_upload_bytes,
        "max_archive_entries": max_archive_entries,
        "object_buckets": buckets,
        "endpoints": {
            "capabilities": "/__ragtime/capabilities",
            "session": "/__ragtime/session",
            "parse_document": "/__ragtime/parse-document",
            "parse_spreadsheet": "/__ragtime/parse-spreadsheet",
            "normalize_table": "/__ragtime/tables/normalize",
            "render_document_preview": "/__ragtime/documents/render-preview",
            "extract_archive": "/__ragtime/archives/extract",
            "upload_target": "/__ragtime/upload-target",
            "files": "/__ragtime/files/{path}",
            "objects": "/__ragtime/objects/{bucket}/{key}",
            "progress": "/__ragtime/progress/{taskId}",
            "jobs": "/__ragtime/jobs/{jobId}",
        },
    }


async def _broadcast_collab_message(
    workspace_id: str,
    file_path: str,
    message: dict[str, Any],
) -> None:
    clients = await _runtime_service().get_collab_clients(workspace_id, file_path)
    payload = json.dumps(message)
    if not clients:
        return

    async def _send(client: WebSocket) -> None:
        with contextlib.suppress(Exception):
            await client.send_text(payload)

    await asyncio.gather(*(_send(client) for client in clients))


def _is_preview_host_request(request: Request) -> bool:
    # Lazy import avoids the import cycle: preview_host imports runtime_routes.
    from ragtime.userspace.preview_host import is_preview_host

    return is_preview_host(request.headers.get("host"))


def _proxy_request_headers(request: Request, *, allow_user_cookies: bool = False) -> dict[str, str]:
    """Build headers to forward to the workspace devserver.

    Sensitive credentials from the *Ragtime* session must not leak to
    the untrusted user-controlled devserver process.

    ``allow_user_cookies`` opts into forwarding the workspace app's own
    (namespaced) cookies. The caller is responsible for ensuring this is only
    ``True`` for requests that arrived on a dedicated per-workspace preview
    host, where app cookies live on an origin isolated from the Ragtime app;
    the sole production caller (``_proxy_http_request``) enforces this via
    ``_is_preview_host_request``.
    """
    _blocked = {
        # Hop-by-hop
        "host",
        "connection",
        "keep-alive",
        "proxy-authenticate",
        "proxy-authorization",
        "te",
        "trailers",
        "transfer-encoding",
        "upgrade",
        "content-length",
        # Credentials - must not reach the untrusted devserver
        "authorization",
        "cookie",
        "x-api-key",
        "x-userspace-share-password",
        "x-userspace-share-auth",
    }
    forwarded_headers = {key: value for key, value in request.headers.items() if key.lower() not in _blocked}
    if allow_user_cookies:
        cookie_header = _sanitize_user_app_cookie_header(request.headers.get("cookie"))
        if cookie_header:
            forwarded_headers["cookie"] = cookie_header
    # Preserve request context for framework URL generation and redirects.
    forwarded_headers.setdefault("x-forwarded-proto", request.url.scheme)
    forwarded_headers.setdefault("x-forwarded-host", request.headers.get("host", ""))
    client_host = request.client.host if request.client else ""
    if client_host:
        forwarded_headers.setdefault("x-forwarded-for", client_host)
    return forwarded_headers


def _proxy_response_headers(
    headers: httpx.Headers,
    *,
    allow_user_cookies: bool = False,
) -> tuple[dict[str, str], list[str]]:
    """Filter devserver response headers.

    ``set-cookie`` is always stripped from the returned header map; when
    ``allow_user_cookies`` is set, the workspace app's namespaced cookies are
    returned separately so the caller can append them as discrete repeated
    ``Set-Cookie`` headers (comma-joining cookies is unsafe). Callers must only
    pass ``allow_user_cookies=True`` for preview-host responses; see
    ``_proxy_http_request`` which gates this on ``_is_preview_host_request``.
    """
    blocked = {
        "connection",
        "keep-alive",
        "proxy-authenticate",
        "proxy-authorization",
        "te",
        "trailers",
        "transfer-encoding",
        "upgrade",
        # Untrusted preview apps must not set first-party cookies on ragtime origin.
        "set-cookie",
        # Devserver frameworks (Express+helmet, Django, etc.) may emit these
        # headers which prevent the content from rendering inside the platform
        # iframe.  The iframe sandbox attribute is the real security boundary;
        # devserver-originated policies are not trustworthy anyway.
        "x-frame-options",
        "content-security-policy",
        "content-security-policy-report-only",
    }
    out = {key: value for key, value in headers.items() if key.lower() not in blocked}
    set_cookies: list[str] = []
    if allow_user_cookies:
        set_cookies = [value for value in headers.get_list("set-cookie") if _is_user_app_set_cookie(value)]
    return out, set_cookies


def _append_set_cookie_headers(response: Response, set_cookie_headers: Sequence[str]) -> Response:
    # Each value is appended as its own ``Set-Cookie`` header line via the
    # public MutableHeaders API; cookies must never be comma-joined into a
    # single header because cookie attribute values can contain commas.
    for value in set_cookie_headers:
        response.headers.append("set-cookie", value)
    return response


def _make_proxy_response_uncacheable(
    headers: dict[str, str],
    *,
    clear_site_cache: bool = False,
) -> None:
    headers.pop("etag", None)
    headers.pop("ETag", None)
    headers.pop("last-modified", None)
    headers.pop("Last-Modified", None)
    headers.pop("expires", None)
    headers.pop("Expires", None)
    headers.pop("age", None)
    headers.pop("Age", None)
    headers["cache-control"] = "no-store, no-cache, must-revalidate, max-age=0"
    headers["pragma"] = "no-cache"
    if clear_site_cache:
        headers["clear-site-data"] = '"cache"'


def _is_html_media_type(media_type: str) -> bool:
    return "text/html" in (media_type or "").lower()


def _is_secure_browser_request(request: Request) -> bool:
    forwarded_proto = request.headers.get("x-forwarded-proto", "").split(",", 1)[0].strip().lower()
    return (
        request.url.scheme == "https"
        or forwarded_proto == "https"
        or request.headers.get("x-forwarded-ssl", "").lower() == "on"
        or request.headers.get("x-scheme", "").lower() == "https"
    )


def _should_clear_site_data_for_proxy_response(request: Request, media_type: str) -> bool:
    return _is_html_media_type(media_type) and _is_secure_browser_request(request)


def _extract_capability_token_from_request(request: Request) -> str | None:
    # Capability tokens must travel via Authorization/explicit header only.
    # Query-string transport is rejected to prevent leakage via browser
    # history, Referer headers, and reverse-proxy access logs.
    auth_header = request.headers.get("authorization", "")
    if auth_header.startswith("Bearer "):
        token = auth_header[7:].strip()
        if token:
            return token
    explicit = request.headers.get("x-userspace-capability-token", "").strip()
    if explicit:
        return explicit
    return None


def _extract_capability_token_from_websocket(websocket: WebSocket) -> str | None:
    # Capability tokens must travel via Authorization/explicit header only.
    # See _extract_capability_token_from_request.
    auth_header = websocket.headers.get("authorization", "")
    if auth_header.startswith("Bearer "):
        token = auth_header[7:].strip()
        if token:
            return token
    explicit = websocket.headers.get("x-userspace-capability-token", "").strip()
    if explicit:
        return explicit
    return None


def _extract_cookie_token_from_request(
    request: Request,
    cookie_name: str,
) -> str | None:
    token = request.cookies.get(cookie_name, "").strip()
    return token or None


def _extract_cookie_token_from_websocket(
    websocket: WebSocket,
    cookie_name: str,
) -> str | None:
    token = websocket.cookies.get(cookie_name, "").strip()
    return token or None


def _require_workspace_capability(
    token: str | None,
    workspace_id: str,
    capability: str,
) -> tuple[dict[str, Any], str]:
    if not token:
        raise HTTPException(status_code=401, detail="Capability token required")
    claims = _runtime_service().verify_capability_token(
        token,
        workspace_id,
        capability,
    )
    user_id = str(claims.get("sub") or "").strip()
    if not user_id:
        raise HTTPException(status_code=401, detail="Capability token missing user")
    return claims, user_id


def _to_websocket_url(http_url: str) -> str:
    parsed = urlsplit(http_url)
    if parsed.scheme == "https":
        scheme = "wss"
    elif parsed.scheme == "http":
        scheme = "ws"
    else:
        scheme = parsed.scheme
    return urlunsplit((scheme, parsed.netloc, parsed.path, parsed.query, parsed.fragment))


def _sanitize_preview_query(query: str | None) -> str | None:
    if not query:
        return None
    internal_query_params = {"cap_token", _PREVIEW_HANDOFF_QUERY_PARAM}
    cleaned_pairs = [(key, value) for key, value in parse_qsl(query, keep_blank_values=True) if key not in internal_query_params]
    if not cleaned_pairs:
        return None
    return urlencode(cleaned_pairs, doseq=True)


def _split_pty_token_from_url(ws_url: str) -> tuple[str, str | None]:
    """Strip the PTY ``token`` query param and return (cleaned_url, token).

    The runtime manager emits PTY WebSocket URLs with the per-session
    token embedded in the query string. Returning it as a separate value
    lets callers forward it via a header so the token never appears in
    worker access logs or any URL-based observability surface.
    """
    parsed = urlsplit(ws_url)
    pairs = parse_qsl(parsed.query, keep_blank_values=True)
    pty_token: str | None = None
    cleaned_pairs: list[tuple[str, str]] = []
    for key, value in pairs:
        if key == "token" and pty_token is None:
            pty_token = value or None
            continue
        cleaned_pairs.append((key, value))
    cleaned_query = urlencode(cleaned_pairs, doseq=True) if cleaned_pairs else ""
    cleaned_url = urlunsplit((parsed.scheme, parsed.netloc, parsed.path, cleaned_query, parsed.fragment))
    return cleaned_url, pty_token


async def _proxy_websocket_request(
    websocket: WebSocket,
    upstream_url: str,
    *,
    read_only: bool = False,
    additional_headers: dict[str, str] | None = None,
) -> None:
    try:
        websockets_module = importlib.import_module("websockets")
    except Exception:
        await _safe_close_websocket(websocket, code=1011)
        return

    requested_subprotocols = [str(protocol).strip() for protocol in (websocket.scope.get("subprotocols") or []) if str(protocol).strip()]

    # Mirror the worker auth token that _proxy_http_request sends so the
    # upstream runtime worker accepts the WebSocket connection.
    extra_headers: dict[str, str] = {}
    worker_token = getattr(settings, "userspace_runtime_worker_auth_token", "") or ""
    if worker_token:
        extra_headers["authorization"] = f"Bearer {worker_token}"
    if additional_headers:
        for header_name, header_value in additional_headers.items():
            if header_value:
                extra_headers[header_name] = header_value

    try:
        async with websockets_module.connect(
            upstream_url,
            max_size=None,
            open_timeout=20,
            subprotocols=requested_subprotocols or None,
            additional_headers=extra_headers or None,
        ) as upstream:
            await websocket.accept(subprotocol=getattr(upstream, "subprotocol", None) or None)

            if read_only:
                await websocket.send_text(
                    json.dumps(
                        {
                            "type": "status",
                            "read_only": True,
                            "message": "Read-only terminal session",
                        }
                    )
                )

            async def downstream_to_upstream() -> None:
                while True:
                    message = await websocket.receive()
                    message_type = message.get("type")
                    if message_type == "websocket.disconnect":
                        break
                    text = message.get("text")
                    data = message.get("bytes")
                    if read_only and text is not None:
                        try:
                            payload = json.loads(text)
                            if payload.get("type") == "input":
                                continue
                        except (json.JSONDecodeError, TypeError):
                            pass
                    if text is not None:
                        await upstream.send(text)
                    elif data is not None:
                        await upstream.send(data)

            async def upstream_to_downstream() -> None:
                while True:
                    message = await upstream.recv()
                    if isinstance(message, bytes):
                        await websocket.send_bytes(message)
                    else:
                        await websocket.send_text(str(message))

            down_task = asyncio.create_task(downstream_to_upstream())
            up_task = asyncio.create_task(upstream_to_downstream())
            done, pending = await asyncio.wait(
                {down_task, up_task},
                return_when=asyncio.FIRST_COMPLETED,
            )
            for task in pending:
                task.cancel()
            for task in done:
                task.result()
    except WebSocketDisconnect:
        return
    except TimeoutError:
        with contextlib.suppress(Exception):
            await websocket.send_text(
                json.dumps(
                    {
                        "type": "status",
                        "read_only": read_only,
                        "message": "Runtime terminal is still starting. Reconnecting...",
                    }
                )
            )
        await _safe_close_websocket(websocket, code=1011)
    except Exception:
        await _safe_close_websocket(websocket, code=1011)


async def _proxy_http_request(
    request: Request,
    upstream_url: str,
    *,
    bridge_workspace_id: str | None = None,
    bridge_context: dict[str, Any] | None = None,
    bridge_script_src: str | None = None,
    primitive_session_factory: Callable[[], Awaitable[dict[str, Any]]] | None = None,
    allow_user_cookies: bool = False,
) -> Response:
    if request.headers.get("upgrade", "").lower() == "websocket":
        raise HTTPException(
            status_code=501,
            detail="WebSocket proxy upgrades are not available on this route",
        )

    # Single authoritative gate for app-cookie forwarding: only honored on
    # dedicated preview hosts, where the workspace app lives on an origin
    # isolated from the Ragtime app. Resolved once here and reused for both the
    # request and response cookie handling so the two can never diverge, and so
    # the legacy same-origin proxy (which never passes allow_user_cookies=True)
    # can never replay untrusted app cookies on the Ragtime origin.
    effective_allow_user_cookies = allow_user_cookies and _is_preview_host_request(request)

    body = await request.body()
    headers = _proxy_request_headers(request, allow_user_cookies=effective_allow_user_cookies)

    # Inject runtime worker auth token for upstream worker requests
    worker_token = getattr(settings, "userspace_runtime_worker_auth_token", "") or ""
    if worker_token:
        headers["authorization"] = f"Bearer {worker_token}"

    # Derive proxy read/write timeout from the maximum tool timeout across
    # all configured tools so the proxy never cuts off a legitimate query.
    # Falls back to 320 s if no tools are configured.
    proxy_read_timeout = _get_max_proxy_timeout()
    timeout = httpx.Timeout(connect=2.0, read=proxy_read_timeout, write=proxy_read_timeout, pool=5.0)
    client = httpx.AsyncClient(timeout=timeout, follow_redirects=False)
    try:
        upstream_request = client.build_request(
            method=request.method,
            url=upstream_url,
            content=body if body else None,
            headers=headers,
        )
        upstream_response = await client.send(upstream_request, stream=True)
    except httpx.RequestError as exc:
        await client.aclose()
        raise HTTPException(
            status_code=502,
            detail=f"Runtime preview upstream unavailable: {exc}",
        ) from exc

    media_type = upstream_response.headers.get("content-type", "")
    resp_headers, set_cookie_headers = _proxy_response_headers(upstream_response.headers, allow_user_cookies=effective_allow_user_cookies)
    resp_headers[_USERSPACE_SURFACE_HEADER] = "preview-proxy"
    resp_headers[_USERSPACE_PREVIEW_PROXY_HEADER] = "true"
    _make_proxy_response_uncacheable(
        resp_headers,
        clear_site_cache=_should_clear_site_data_for_proxy_response(request, media_type),
    )

    if _is_html_media_type(media_type):
        try:
            content = await upstream_response.aread()
        finally:
            await upstream_response.aclose()
            await client.aclose()

        sandbox_flags: list[str] | None = None
        try:
            app_settings = await get_app_settings()
            sandbox_flags = list(app_settings.get("userspace_preview_sandbox_flags") or [])
        except Exception:
            sandbox_flags = None
        primitive_session: dict[str, Any] | None = None
        auth_requirement = _extract_ragtime_auth_requirement(content)
        if auth_requirement is not None and primitive_session_factory is not None:
            primitive_session = await primitive_session_factory()
            auth_payload = primitive_session.get("auth") if isinstance(primitive_session, dict) else None
            is_authenticated = bool(auth_payload.get("authenticated")) if isinstance(auth_payload, dict) else False
            if not is_authenticated:
                return RedirectResponse(
                    url=_interactive_auth_url_from_requirement(auth_requirement, _return_to_from_request(request)),
                    status_code=302,
                    headers={
                        "Cache-Control": "no-store",
                        _USERSPACE_SURFACE_HEADER: "preview-proxy",
                        _USERSPACE_PREVIEW_PROXY_HEADER: "true",
                    },
                )
        content = _inject_bridge_script(
            content,
            sandbox_flags,
            workspace_id=bridge_workspace_id,
            bridge_context=bridge_context,
            bridge_script_src=bridge_script_src,
            primitive_session=primitive_session,
            cleanup_preview_handoff=_PREVIEW_HANDOFF_QUERY_PARAM in request.query_params,
        )
        # The response body may be decoded by httpx and is definitely mutated by
        # bridge injection, so upstream encoding and length metadata is stale.
        resp_headers.pop("content-encoding", None)
        resp_headers.pop("content-length", None)

        return _append_set_cookie_headers(
            Response(
                content=content,
                status_code=upstream_response.status_code,
                headers=resp_headers,
                media_type=media_type or None,
            ),
            set_cookie_headers,
        )

    async def _iter_stream() -> AsyncIterator[bytes]:
        try:
            async for chunk in upstream_response.aiter_raw():
                yield chunk
        except asyncio.CancelledError:
            return
        finally:
            await upstream_response.aclose()
            await client.aclose()

    return _append_set_cookie_headers(
        _CancellationSafeStreamingResponse(
            _iter_stream(),
            status_code=upstream_response.status_code,
            headers=resp_headers,
            media_type=media_type or None,
        ),
        set_cookie_headers,
    )


_BRIDGE_CONFIG_MARKER = b"__ragtime_preview_sandbox_flags"
_BRIDGE_CONTEXT_MARKER = b"__ragtime_preview_bridge"
_PRIMITIVE_SESSION_MARKER = b"__ragtime_session"
_PREVIEW_HANDOFF_QUERY_PARAM = "__ragtime_preview_handoff"
_PREVIEW_HANDOFF_CLEANUP_MARKER = b"__ragtime_cleanup_preview_handoff"
_BRIDGE_DETECT_RE = re.compile(rb"bridge\.js", re.IGNORECASE)
_HEAD_CLOSE_RE = re.compile(rb"(</head\s*>)", re.IGNORECASE)
_FIRST_SCRIPT_RE = re.compile(rb"(<script[\s>])", re.IGNORECASE)
_RAGTIME_AUTH_META_RE = re.compile(
    rb"<meta\s+[^>]*(?:name\s*=\s*['\"]ragtime-auth['\"][^>]*content\s*=\s*['\"]([^'\"]*)['\"]|content\s*=\s*['\"]([^'\"]*)['\"][^>]*name\s*=\s*['\"]ragtime-auth['\"])[^>]*>",
    re.IGNORECASE,
)


def _json_for_inline_script(value: Any) -> bytes:
    serialized = json.dumps(value, separators=(",", ":"), default=str)
    serialized = serialized.replace("<", "\\u003c").replace(">", "\\u003e").replace("&", "\\u0026")
    return serialized.encode("utf-8")


def _parse_ragtime_auth_meta_content(raw_content: bytes) -> dict[str, Any] | None:
    try:
        content = raw_content.decode("utf-8", errors="ignore")
    except Exception:
        return None
    values: dict[str, str] = {}
    required = False
    for item in re.split(r"[;\n]+", content):
        part = item.strip()
        if not part:
            continue
        if "=" in part:
            key, value = part.split("=", 1)
            values[key.strip().lower()] = value.strip().strip("\"'")
            continue
        if part.lower() in {"required", "require", "auth_required"}:
            required = True
    if str(values.get("required") or "").strip().lower() in {"1", "true", "yes", "required"}:
        required = True
    if not required:
        return None

    raw_surfaces = values.get("surfaces") or values.get("surface") or ""
    surfaces: list[UserSpaceBrowserSurface] = []
    for part in re.split(r"[,\s]+", raw_surfaces):
        surface = part.strip()
        if surface in {"collab", "runtime_pty"}:
            surfaces.append(cast(UserSpaceBrowserSurface, surface))
    return {
        "required": True,
        "surfaces": _normalize_browser_surfaces(surfaces or None),
        "auth_method_key": normalize_auth_method_key(values.get("auth_method_key") or values.get("method") or None) or None,
    }


def _extract_ragtime_auth_requirement(html: bytes) -> dict[str, Any] | None:
    match = _RAGTIME_AUTH_META_RE.search(html)
    if not match:
        return None
    raw_content = match.group(1) or match.group(2) or b""
    return _parse_ragtime_auth_meta_content(raw_content)


def _return_to_from_request(request: Request) -> str:
    path = request.url.path or "/"
    if not path.startswith("/") or path.startswith("//"):
        path = "/"
    query = request.url.query
    return f"{path}?{query}" if query else path


def _interactive_auth_url_from_requirement(requirement: dict[str, Any], return_to: str) -> str:
    params: list[tuple[str, str]] = []
    for surface in _normalize_browser_surfaces(requirement.get("surfaces") or None):
        params.append(("surfaces", surface))
    auth_method_key = normalize_auth_method_key(requirement.get("auth_method_key"))
    if auth_method_key:
        params.append(("auth_method_key", auth_method_key))
    params.append(("return_to", return_to if return_to.startswith("/") and not return_to.startswith("//") else "/"))
    return "/__ragtime/browser-auth/start?" + urlencode(params)


def _build_primitive_session_tag(primitive_session: dict[str, Any]) -> bytes:
    serialized_session = _json_for_inline_script(primitive_session)
    return (
        b"<script>window.__ragtime_session=" + serialized_session + b";window.__ragtime_auth=window.__ragtime_session&&window.__ragtime_session.auth;</script>"
    )


def _build_preview_handoff_cleanup_tag() -> bytes:
    param = json.dumps(_PREVIEW_HANDOFF_QUERY_PARAM).encode("utf-8")
    return (
        b"<script>window.__ragtime_cleanup_preview_handoff=true;(function(){try{var url=new URL(window.location.href);"
        b"var param="
        + param
        + b";if(url.searchParams.has(param)){url.searchParams.delete(param);window.history.replaceState(window.history.state,'',url.pathname+url.search+url.hash);}}catch(err){}})();</script>"
    )


def _build_bridge_script_tag(
    workspace_id: str | None = None,
    *,
    bridge_script_src: str | None = None,
) -> bytes:
    params = {"workspace_id": workspace_id} if workspace_id else {}
    query = urlencode(params)
    src = (bridge_script_src or "/__ragtime/bridge.js").strip() or "/__ragtime/bridge.js"
    if query:
        joiner = "&" if "?" in src else "?"
        src = f"{src}{joiner}{query}"
    return f'<script src="{src}"></script>'.encode("utf-8")


def _build_bridge_config_tag(sandbox_flags: list[str]) -> bytes:
    normalized_flags = [flag.strip() for flag in sandbox_flags if isinstance(flag, str) and flag.strip()]
    serialized_flags = _json_for_inline_script(normalized_flags)
    return b"<script>window.__ragtime_preview_sandbox_flags=" + serialized_flags + b";</script>"


def _build_bridge_context_tag(bridge_context: dict[str, Any]) -> bytes:
    serialized_context = _json_for_inline_script(bridge_context)
    return b"<script>window.__ragtime_preview_bridge=" + serialized_context + b";</script>"


def _inject_bridge_script(
    html: bytes,
    sandbox_flags: list[str] | None = None,
    *,
    workspace_id: str | None = None,
    bridge_context: dict[str, Any] | None = None,
    bridge_script_src: str | None = None,
    primitive_session: dict[str, Any] | None = None,
    cleanup_preview_handoff: bool = False,
) -> bytes:
    """Inject the platform data-bridge script into HTML responses.

    Inserts ``<script src="/__ragtime/bridge.js?...">`` before
    ``</head>`` or the first ``<script`` tag so ``window.__ragtime_context``
    and platform visualization libraries are available to workspace code.
    Skips injection if bridge.js is already referenced.
    """
    injected = b""
    if cleanup_preview_handoff and _PREVIEW_HANDOFF_CLEANUP_MARKER not in html:
        injected += _build_preview_handoff_cleanup_tag() + b"\n"
    if sandbox_flags is not None and _BRIDGE_CONFIG_MARKER not in html:
        injected += _build_bridge_config_tag(sandbox_flags) + b"\n"
    if bridge_context is not None and _BRIDGE_CONTEXT_MARKER not in html:
        injected += _build_bridge_context_tag(bridge_context) + b"\n"
    if primitive_session is not None and _PRIMITIVE_SESSION_MARKER not in html:
        injected += _build_primitive_session_tag(primitive_session) + b"\n"
    if not _BRIDGE_DETECT_RE.search(html):
        injected += (
            _build_bridge_script_tag(
                workspace_id,
                bridge_script_src=bridge_script_src,
            )
            + b"\n"
        )
    if not injected:
        return html
    head_match = _HEAD_CLOSE_RE.search(html)
    script_match = _FIRST_SCRIPT_RE.search(html)
    candidates = [match for match in (head_match, script_match) if match is not None]
    if candidates:
        insert_at = min(match.start() for match in candidates)
        return html[:insert_at] + injected + html[insert_at:]
    return injected + html


@router.get(
    "/runtime/workspaces/{workspace_id}/session",
    response_model=UserSpaceRuntimeSessionResponse,
)
async def get_runtime_session(
    workspace_id: str,
    user: Any = Depends(get_current_user),
):
    return await _runtime_service().get_runtime_session(workspace_id, user.id)


@router.post(
    "/runtime/workspaces/{workspace_id}/session/start",
    response_model=UserSpaceRuntimeActionResponse,
)
async def start_runtime_session(
    workspace_id: str,
    user: Any = Depends(get_current_user),
):
    return await _runtime_service().start_runtime_session(workspace_id, user.id)


@router.post(
    "/runtime/workspaces/{workspace_id}/session/stop",
    response_model=UserSpaceRuntimeActionResponse,
)
async def stop_runtime_session(
    workspace_id: str,
    user: Any = Depends(get_current_user),
):
    return await _runtime_service().stop_runtime_session(workspace_id, user.id)


@router.get(
    "/runtime/workspaces/{workspace_id}/events",
)
async def workspace_events_sse(
    workspace_id: str,
    request: Request,
    _user: Any = Depends(get_current_user),
):
    """SSE stream that emits a message whenever the workspace generation advances."""

    await _runtime_service().track_workspace_runtime_events(workspace_id)

    async def _stream():
        # Always emit the current generation so the client knows the
        # connection is live and what the baseline is.
        try:
            generation = await _runtime_service().get_workspace_generation(workspace_id)
            initial_payload = _runtime_service().get_workspace_event_payload(workspace_id)
            yield f"data: {json.dumps(initial_payload)}\n\n"

            while True:
                if await request.is_disconnected():
                    break
                new_gen = await _runtime_service().wait_workspace_generation(workspace_id, generation, timeout=25.0)
                if await request.is_disconnected():
                    break
                if new_gen > generation:
                    generation = new_gen
                    payload = _runtime_service().get_workspace_event_payload(workspace_id)
                    yield f"data: {json.dumps(payload)}\n\n"
                else:
                    # Keepalive – SSE comment to prevent proxy/browser timeouts
                    yield ": keepalive\n\n"
        except asyncio.CancelledError:
            return

    return _CancellationSafeStreamingResponse(
        _stream(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
        },
    )


@router.get(
    "/runtime/workspaces/{workspace_id}/devserver/status",
    response_model=UserSpaceRuntimeStatusResponse,
)
async def get_devserver_status(
    workspace_id: str,
    request: Request,
    user: Any = Depends(get_current_user),
):
    status = await _runtime_service().get_devserver_status(workspace_id, user.id)
    status.preview_url = _runtime_service().get_preview_origin(
        workspace_id,
        control_plane_origin=get_browser_matched_origin(request),
    )
    return status


@router.get(
    "/runtime/workspaces/{workspace_id}/tab-state",
    response_model=UserSpaceWorkspaceTabStateResponse,
)
async def get_workspace_tab_state(
    workspace_id: str,
    request: Request,
    selected_conversation_id: str | None = None,
    user: Any = Depends(get_current_user),
):
    tab_state = await _runtime_service().get_workspace_tab_state(
        workspace_id,
        user.id,
        selected_conversation_id=selected_conversation_id,
        is_admin=bool(getattr(user, "role", None) == "admin"),
    )
    tab_state.runtime_status.preview_url = _runtime_service().get_preview_origin(
        workspace_id,
        control_plane_origin=get_browser_matched_origin(request),
    )
    return tab_state


@router.post(
    "/runtime/workspaces/{workspace_id}/devserver/restart",
    response_model=UserSpaceRuntimeActionResponse,
)
async def restart_devserver(
    workspace_id: str,
    user: Any = Depends(get_current_user),
):
    return await _runtime_service().restart_devserver(workspace_id, user.id)


@router.post(
    "/runtime/workspaces/{workspace_id}/mounts/{mount_id}/refresh",
    response_model=UserSpaceRuntimeActionResponse,
)
async def refresh_runtime_mount(
    workspace_id: str,
    mount_id: str,
    user: Any = Depends(get_current_user),
):
    result = await _runtime_service().refresh_workspace_mount(
        workspace_id,
        user.id,
        mount_id,
    )
    await _runtime_service().bump_workspace_generation(
        workspace_id,
        "mount_refresh",
    )
    return result


@router.get("/runtime/workspaces/{workspace_id}/screenshots/{filename}")
async def get_runtime_screenshot(
    workspace_id: str,
    filename: str,
    user: Any = Depends(get_current_user),
):
    await _userspace_service().enforce_workspace_role(workspace_id, user.id, "viewer")

    normalized_name = (filename or "").strip().replace("\\", "/")
    basename = Path(normalized_name).name
    if not basename or basename != normalized_name or not basename.lower().endswith(".png"):
        raise HTTPException(status_code=400, detail="Invalid screenshot filename")

    index_data_root = Path(os.getenv("INDEX_DATA_PATH", "/data"))
    screenshot_dir = (index_data_root / "_userspace" / "workspaces" / workspace_id / "runtime-artifacts" / "screenshots").resolve()
    screenshot_path = (screenshot_dir / basename).resolve()

    try:
        screenshot_path.relative_to(screenshot_dir)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid screenshot path") from exc

    if not screenshot_path.exists() or not screenshot_path.is_file():
        raise HTTPException(status_code=404, detail="Screenshot not found")

    return FileResponse(
        path=str(screenshot_path),
        media_type="image/png",
        filename=basename,
    )


@router.post(
    "/runtime/workspaces/{workspace_id}/capability-token",
    response_model=UserSpaceCapabilityTokenResponse,
)
async def issue_capability_token(
    workspace_id: str,
    payload: dict[str, Any],
    user: Any = Depends(get_current_user),
):
    capabilities = payload.get("capabilities") if isinstance(payload, dict) else []
    if not isinstance(capabilities, list):
        capabilities = []
    session_id = payload.get("session_id") if isinstance(payload, dict) else None
    return await _runtime_service().issue_capability_token(
        workspace_id,
        user.id,
        [str(value) for value in capabilities],
        str(session_id) if session_id else None,
    )


@router.post(
    "/runtime/workspaces/{workspace_id}/browser-auth",
    response_model=UserSpaceBrowserAuthResponse,
)
async def authorize_browser_surfaces(
    workspace_id: str,
    payload: UserSpaceBrowserAuthRequest,
    request: Request,
    response: Response,
    user: Any = Depends(get_current_user),
):
    return await _authorize_browser_surfaces_for_user_id(
        workspace_id,
        payload,
        request,
        response,
        user_id=user.id,
    )


async def _authorize_browser_surfaces_for_user_id(
    workspace_id: str,
    payload: UserSpaceBrowserAuthRequest,
    request: Request,
    response: Response,
    *,
    user_id: str,
) -> UserSpaceBrowserAuthResponse:
    surfaces = _normalize_browser_surfaces(payload.surfaces)
    auth_method_key = await _validate_auth_binding_method(payload.auth_method_key)
    authorizations: list[UserSpaceBrowserAuthorization] = []

    for surface in surfaces:
        config = _BROWSER_SURFACE_COOKIE_CONFIG[surface]
        token_response = await _runtime_service().issue_capability_token(
            workspace_id,
            user_id,
            list(config["capabilities"]),
        )
        max_age = max(
            60,
            int((token_response.expires_at - datetime.now(timezone.utc)).total_seconds()),
        )
        response.set_cookie(
            key=str(config["cookie_name"]),
            value=token_response.token,
            max_age=max_age,
            httponly=True,
            secure=request.url.scheme == "https",
            samesite="lax",
            path=str(config["path_builder"](workspace_id)),
        )
        authorizations.append(
            UserSpaceBrowserAuthorization(
                surface=surface,
                expires_at=token_response.expires_at,
                auth_method_key=auth_method_key,
            )
        )

    return UserSpaceBrowserAuthResponse(
        workspace_id=workspace_id,
        auth_method_key=auth_method_key,
        authorizations=authorizations,
    )


def _clear_browser_surface_cookies(
    workspace_id: str,
    response: Response,
) -> None:
    for config in _BROWSER_SURFACE_COOKIE_CONFIG.values():
        response.delete_cookie(
            key=str(config["cookie_name"]),
            path=str(config["path_builder"](workspace_id)),
        )


@router.post("/runtime/workspaces/{workspace_id}/browser-auth/logout")
async def logout_browser_surfaces(
    workspace_id: str,
    response: Response,
    user: Any = Depends(get_current_user),
):
    await _userspace_service().enforce_workspace_role(workspace_id, user.id, "viewer")
    _clear_browser_surface_cookies(workspace_id, response)
    return {"success": True}


@router.post("/runtime/workspaces/{workspace_id}/documents/parse")
async def runtime_parse_document(
    workspace_id: str,
    file: UploadFile = File(...),
    user: Any = Depends(get_current_user),
):
    await _userspace_service().enforce_workspace_role(workspace_id, user.id, "viewer")
    return await _parse_uploaded_document_upload(file)


@router.post("/runtime/workspaces/{workspace_id}/spreadsheets/parse")
async def runtime_parse_spreadsheet(
    workspace_id: str,
    file: UploadFile = File(...),
    user: Any = Depends(get_current_user),
):
    await _userspace_service().enforce_workspace_role(workspace_id, user.id, "viewer")
    return await _parse_uploaded_spreadsheet_upload(file)


@router.get("/runtime/workspaces/{workspace_id}/primitives/capabilities")
async def runtime_primitive_capabilities(
    workspace_id: str,
    user: Any = Depends(get_current_user),
):
    return await _primitive_capabilities(workspace_id, user.id, preview_mode="workspace")


@router.get("/runtime/workspaces/{workspace_id}/primitives/session")
async def runtime_primitive_session(
    workspace_id: str,
    user: Any = Depends(get_current_user),
):
    return await _primitive_session_payload(workspace_id, user.id, mode="workspace", user=user)


@router.post("/runtime/workspaces/{workspace_id}/primitives/tables/normalize")
async def runtime_primitive_table_normalize(
    workspace_id: str,
    file: UploadFile = File(...),
    user: Any = Depends(get_current_user),
):
    await _userspace_service().enforce_workspace_role(workspace_id, user.id, "viewer")
    return await _normalize_uploaded_table_upload(file)


@router.post("/runtime/workspaces/{workspace_id}/primitives/documents/render-preview")
async def runtime_primitive_document_render_preview(
    workspace_id: str,
    file: UploadFile = File(...),
    user: Any = Depends(get_current_user),
):
    await _userspace_service().enforce_workspace_role(workspace_id, user.id, "viewer")
    return await _render_uploaded_document_preview_upload(file)


@router.post("/runtime/workspaces/{workspace_id}/primitives/archives/extract")
async def runtime_primitive_archive_extract(
    workspace_id: str,
    request: Request,
    target: str = "files",
    destination_path: str = "",
    bucket: str | None = None,
    user: Any = Depends(get_current_user),
):
    return await _primitive_archive_extract(
        workspace_id,
        request,
        user.id,
        target=target,
        destination_path=destination_path,
        bucket_name=bucket,
    )


@router.post("/runtime/workspaces/{workspace_id}/primitives/upload-target")
async def runtime_primitive_upload_target(
    workspace_id: str,
    payload: dict[str, Any],
    user: Any = Depends(get_current_user),
):
    return await _primitive_upload_target(workspace_id, payload, user.id, preview_origin=False)


@router.get("/runtime/workspaces/{workspace_id}/primitives/files/{file_path:path}")
async def runtime_primitive_file_read(
    workspace_id: str,
    file_path: str,
    user: Any = Depends(get_current_user),
):
    return await _primitive_workspace_file_response(workspace_id, file_path, user.id)


@router.put("/runtime/workspaces/{workspace_id}/primitives/files/{file_path:path}")
async def runtime_primitive_file_write(
    workspace_id: str,
    file_path: str,
    request: Request,
    user: Any = Depends(get_current_user),
):
    return await _primitive_workspace_file_write_request(workspace_id, file_path, request, user.id)


@router.get("/runtime/workspaces/{workspace_id}/primitives/objects/{bucket_name}/{object_path:path}")
async def runtime_primitive_object_read(
    workspace_id: str,
    bucket_name: str,
    object_path: str,
    user: Any = Depends(get_current_user),
):
    return await _primitive_object_response(workspace_id, bucket_name, object_path, user.id)


@router.put("/runtime/workspaces/{workspace_id}/primitives/objects/{bucket_name}/{object_path:path}")
async def runtime_primitive_object_write(
    workspace_id: str,
    bucket_name: str,
    object_path: str,
    request: Request,
    user: Any = Depends(get_current_user),
):
    return await _primitive_object_write_request(workspace_id, bucket_name, object_path, request, user.id)


@router.get("/runtime/workspaces/{workspace_id}/primitives/progress/{task_id}")
async def runtime_primitive_progress_get(
    workspace_id: str,
    task_id: str,
    user: Any = Depends(get_current_user),
):
    return await _primitive_progress_get(workspace_id, task_id, user.id)


@router.put("/runtime/workspaces/{workspace_id}/primitives/progress/{task_id}")
async def runtime_primitive_progress_put(
    workspace_id: str,
    task_id: str,
    payload: dict[str, Any],
    user: Any = Depends(get_current_user),
):
    return await _primitive_progress_put(workspace_id, task_id, payload, user.id)


@router.post("/runtime/workspaces/{workspace_id}/primitives/jobs")
async def runtime_primitive_job_create(
    workspace_id: str,
    payload: dict[str, Any],
    user: Any = Depends(get_current_user),
):
    return await _primitive_job_create(workspace_id, payload, user.id)


@router.get("/runtime/workspaces/{workspace_id}/primitives/jobs/{job_id}")
async def runtime_primitive_job_get(
    workspace_id: str,
    job_id: str,
    user: Any = Depends(get_current_user),
):
    return await _primitive_job_get(workspace_id, job_id, user.id)


@router.put("/runtime/workspaces/{workspace_id}/primitives/jobs/{job_id}")
async def runtime_primitive_job_put(
    workspace_id: str,
    job_id: str,
    payload: dict[str, Any],
    user: Any = Depends(get_current_user),
):
    return await _primitive_job_put(workspace_id, job_id, payload, user.id)


@router.post(
    "/runtime/workspaces/{workspace_id}/preview-launch",
    response_model=UserSpacePreviewLaunchResponse,
)
async def issue_workspace_preview_launch(
    workspace_id: str,
    payload: UserSpacePreviewLaunchRequest,
    request: Request,
    user: Any = Depends(get_current_user),
    session_token: str | None = Depends(get_session_token),
):
    external_origin = get_browser_matched_origin(
        request,
        browser_origin=payload.parent_origin,
    )
    token_data = decode_access_token(session_token) if session_token else None
    preview_origin, expires_at, warning = await _runtime_service().describe_workspace_preview_launch(
        workspace_id,
        user.id,
        control_plane_origin=external_origin,
        session_expires_at=token_data.exp if token_data else None,
    )
    if warning and warning.issue_code == "preview_host_unreachable":
        warning = None
    return _workspace_preview_entry_launch_response(
        workspace_id=workspace_id,
        base_url=external_origin,
        target_path=payload.path,
        parent_origin=payload.parent_origin,
        preview_origin=preview_origin,
        expires_at=expires_at,
        preview_warning=warning,
    )


@router.get(
    "/runtime/workspaces/{workspace_id}/preview-entry",
)
async def workspace_preview_entry(
    workspace_id: str,
    request: Request,
    path: str = "/",
    parent_origin: str | None = None,
    user: Any = Depends(get_current_user),
):
    launch = await _runtime_service().issue_workspace_preview_launch(
        workspace_id,
        user.id,
        control_plane_origin=get_browser_matched_origin(
            request,
            browser_origin=parent_origin,
        ),
        path=path,
        parent_origin=parent_origin,
    )
    warning = getattr(launch, "preview_warning", None)
    if warning and warning.issue_code == "preview_host_unreachable":
        return _preview_host_unreachable_response(
            workspace_id=workspace_id,
            preview_origin=launch.preview_origin,
            warning=warning,
        )
    return Response(
        status_code=307,
        headers={
            "Location": launch.preview_url,
            "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
            "Pragma": "no-cache",
            "Referrer-Policy": "no-referrer",
        },
    )


@router.post(
    "/shared/{share_token}/preview-launch",
    response_model=UserSpacePreviewLaunchResponse,
)
@limiter.limit(SHARE_AUTH_RATE_LIMIT)
async def issue_shared_preview_launch(
    share_token: str,
    payload: UserSpacePreviewLaunchRequest,
    request: Request,
    response: Response,
    share_password: str | None = Header(
        default=None,
        alias="X-UserSpace-Share-Password",
    ),
    user: Any | None = Depends(get_current_user_optional),
):
    share_auth_token = share_auth_token_from_request(
        request.headers,
        request.cookies,
        share_token=share_token,
    )
    authorization = await _userspace_service().authorize_shared_workspace_access(
        share_token,
        current_user=user,
        password=share_password,
        share_auth_token=share_auth_token,
    )
    workspace_id = authorization["workspace_id"]
    expires_at = authorization["expires_at"]
    if authorization["share_auth_token"] and expires_at is not None:
        max_age = max(
            60,
            int((expires_at - datetime.now(timezone.utc)).total_seconds()),
        )
        set_share_auth_cookie(
            response,
            authorization["share_auth_token"],
            max_age=max_age,
            secure=request.url.scheme == "https" or bool(getattr(settings, "session_cookie_secure", False)),
            share_token=share_token,
        )
    external_origin = get_browser_matched_origin(
        request,
        browser_origin=payload.parent_origin,
    )
    if payload.prefer_root_proxy:
        return _root_share_launch_response(
            workspace_id=workspace_id,
            base_url=external_origin,
            share_path=f"/shared/{quote(share_token, safe='')}",
            target_path=payload.path,
        )
    share_access_mode = await _userspace_service().get_share_access_mode(workspace_id)
    return await _runtime_service().issue_shared_preview_launch(
        workspace_id,
        control_plane_origin=external_origin,
        path=payload.path,
        parent_origin=payload.parent_origin,
        share_token=share_token,
        subject_user_id=str(getattr(user, "id", "") or "") or None,
        share_access_mode=share_access_mode,
    )


@router.post(
    "/shared/{owner_username}/{share_slug}/preview-launch",
    response_model=UserSpacePreviewLaunchResponse,
)
@limiter.limit(SHARE_AUTH_RATE_LIMIT)
async def issue_shared_preview_launch_by_slug(
    owner_username: str,
    share_slug: str,
    payload: UserSpacePreviewLaunchRequest,
    request: Request,
    response: Response,
    share_password: str | None = Header(
        default=None,
        alias="X-UserSpace-Share-Password",
    ),
    user: Any | None = Depends(get_current_user_optional),
):
    share_auth_token = share_auth_token_from_request(
        request.headers,
        request.cookies,
        owner_username=owner_username,
        share_slug=share_slug,
    )
    authorization = await _userspace_service().authorize_shared_workspace_access_by_slug(
        owner_username,
        share_slug,
        current_user=user,
        password=share_password,
        share_auth_token=share_auth_token,
    )
    workspace_id = authorization["workspace_id"]
    expires_at = authorization["expires_at"]
    if authorization["share_auth_token"] and expires_at is not None:
        max_age = max(
            60,
            int((expires_at - datetime.now(timezone.utc)).total_seconds()),
        )
        set_share_auth_cookie(
            response,
            authorization["share_auth_token"],
            max_age=max_age,
            secure=request.url.scheme == "https" or bool(getattr(settings, "session_cookie_secure", False)),
            owner_username=owner_username,
            share_slug=share_slug,
        )
    external_origin = get_browser_matched_origin(
        request,
        browser_origin=payload.parent_origin,
    )
    if payload.prefer_root_proxy:
        return _root_share_launch_response(
            workspace_id=workspace_id,
            base_url=external_origin,
            share_path=(f"/{quote(owner_username, safe='')}/{quote(share_slug, safe='')}"),
            target_path=payload.path,
        )
    share_access_mode = await _userspace_service().get_share_access_mode(workspace_id)
    return await _runtime_service().issue_shared_preview_launch(
        workspace_id,
        control_plane_origin=external_origin,
        path=payload.path,
        parent_origin=payload.parent_origin,
        owner_username=owner_username,
        share_slug=share_slug,
        subject_user_id=str(getattr(user, "id", "") or "") or None,
        share_access_mode=share_access_mode,
    )


@router.get(
    "/runtime/workspaces/{workspace_id}/fs/{file_path:path}",
    response_model=UserSpaceFileResponse,
)
async def runtime_fs_read(
    workspace_id: str,
    file_path: str,
    request: Request,
):
    token = _extract_capability_token_from_request(request)
    _, user_id = _require_workspace_capability(
        token,
        workspace_id,
        "userspace.runtime_fs_read",
    )
    return await _runtime_service().runtime_fs_read(
        workspace_id,
        file_path,
        user_id,
    )


@router.put(
    "/runtime/workspaces/{workspace_id}/fs/{file_path:path}",
    response_model=UserSpaceFileResponse,
)
async def runtime_fs_write(
    workspace_id: str,
    file_path: str,
    payload: dict[str, Any],
    request: Request,
):
    token = _extract_capability_token_from_request(request)
    _, user_id = _require_workspace_capability(
        token,
        workspace_id,
        "userspace.runtime_fs_write",
    )
    return await _runtime_service().runtime_fs_write(
        workspace_id,
        file_path,
        str(payload.get("content", "")),
        user_id,
    )


@router.delete("/runtime/workspaces/{workspace_id}/fs/{file_path:path}")
async def runtime_fs_delete(
    workspace_id: str,
    file_path: str,
    request: Request,
):
    token = _extract_capability_token_from_request(request)
    _, user_id = _require_workspace_capability(
        token,
        workspace_id,
        "userspace.runtime_fs_write",
    )
    return await _runtime_service().runtime_fs_delete(
        workspace_id,
        file_path,
        user_id,
    )


@router.websocket("/runtime/workspaces/{workspace_id}/pty")
async def runtime_pty(workspace_id: str, websocket: WebSocket):
    token = _extract_cookie_token_from_websocket(
        websocket,
        _RUNTIME_PTY_CAPABILITY_COOKIE,
    )
    try:
        _, user_id = _require_workspace_capability(
            token,
            workspace_id,
            "userspace.runtime_pty",
        )
    except HTTPException:
        await websocket.close(code=4401)
        return

    try:
        await _userspace_service().enforce_workspace_role(workspace_id, user_id, "viewer")
    except HTTPException:
        await websocket.close(code=4403)
        return

    can_write = True
    try:
        await _userspace_service().enforce_workspace_role(workspace_id, user_id, "editor")
    except HTTPException:
        can_write = False

    upstream_ws_url = await _runtime_service().build_workspace_pty_upstream_ws_url(
        workspace_id,
        user_id,
    )
    # The runtime manager mints PTY URLs with the per-session token in the
    # URL query (legacy contract). Move it into an X-PTY-Token header before
    # connecting upstream so the token does not appear in worker access logs
    # or any URL-based observability surface.
    upstream_ws_url, pty_token = _split_pty_token_from_url(upstream_ws_url)
    pty_headers = {"x-pty-token": pty_token} if pty_token else None
    if not can_write:
        await _proxy_websocket_request(
            websocket,
            upstream_ws_url,
            read_only=True,
            additional_headers=pty_headers,
        )
        return

    await _proxy_websocket_request(
        websocket,
        upstream_ws_url,
        additional_headers=pty_headers,
    )


@router.websocket("/collab/workspaces/{workspace_id}/files/{file_path:path}")
async def collab_file_socket(workspace_id: str, file_path: str, websocket: WebSocket):
    accepted = False

    async def _reject_collab(code: int) -> None:
        nonlocal accepted
        if not accepted:
            with contextlib.suppress(Exception):
                await websocket.accept()
            accepted = True
        await _safe_close_websocket(websocket, code)

    token = _extract_cookie_token_from_websocket(
        websocket,
        _COLLAB_CAPABILITY_COOKIE,
    )
    try:
        _, user_id = _require_workspace_capability(
            token,
            workspace_id,
            "userspace.collab_connect",
        )
    except HTTPException:
        await _reject_collab(4401)
        return

    try:
        snapshot = await _runtime_service().get_collab_snapshot(
            workspace_id,
            file_path,
            user_id,
        )
    except HTTPException as exc:
        await _reject_collab(4403 if exc.status_code == 403 else 4404)
        return

    can_edit = not snapshot.read_only

    await websocket.accept()
    accepted = True
    await _runtime_service().register_collab_client(workspace_id, snapshot.file_path, websocket, user_id)

    async def _cleanup_collab() -> None:
        users = await _runtime_service().clear_collab_presence(
            workspace_id,
            snapshot.file_path,
            user_id,
        )
        await _runtime_service().unregister_collab_client(
            workspace_id,
            snapshot.file_path,
            websocket,
        )
        await _broadcast_collab_message(
            workspace_id,
            snapshot.file_path,
            {
                "type": "presence",
                "workspace_id": workspace_id,
                "file_path": snapshot.file_path,
                "users": users,
            },
        )

    try:
        await websocket.send_text(
            json.dumps(
                {
                    "type": "snapshot",
                    "workspace_id": workspace_id,
                    "file_path": snapshot.file_path,
                    "version": snapshot.version,
                    "content": snapshot.content,
                    "read_only": snapshot.read_only,
                }
            )
        )
        current_presence = await _runtime_service().get_collab_presence(
            workspace_id,
            snapshot.file_path,
        )
        await websocket.send_text(
            json.dumps(
                {
                    "type": "presence",
                    "workspace_id": workspace_id,
                    "file_path": snapshot.file_path,
                    "users": current_presence,
                }
            )
        )

        while True:
            data = await websocket.receive_text()
            payload = json.loads(data)
            message_type = payload.get("type")
            if message_type == "presence":
                users = await _runtime_service().update_collab_presence(
                    workspace_id,
                    snapshot.file_path,
                    user_id,
                    payload if isinstance(payload, dict) else {},
                )
                await _broadcast_collab_message(
                    workspace_id,
                    snapshot.file_path,
                    {
                        "type": "presence",
                        "workspace_id": workspace_id,
                        "file_path": snapshot.file_path,
                        "users": users,
                    },
                )
                continue
            if message_type != "update":
                continue
            if not can_edit:
                await websocket.send_text(json.dumps({"type": "error", "message": "Read-only collaboration session"}))
                continue

            content = str(payload.get("content", ""))
            client_version_raw = payload.get("version")
            client_version: int | None = None
            if isinstance(client_version_raw, (int, float, str)):
                version_str = str(client_version_raw).strip()
                if version_str.isdigit():
                    parsed = int(version_str)
                    if parsed > 0:
                        client_version = parsed
            try:
                updated = await _runtime_service().apply_collab_update(
                    workspace_id,
                    snapshot.file_path,
                    content,
                    user_id,
                    expected_version=client_version,
                )
            except RuntimeVersionConflictError as conflict:
                latest = await _runtime_service().get_collab_snapshot(
                    workspace_id,
                    snapshot.file_path,
                    user_id,
                )
                await websocket.send_text(
                    json.dumps(
                        {
                            "type": "error",
                            "message": (f"Version conflict: expected {conflict.expected_version}, current {conflict.actual_version}"),
                        }
                    )
                )
                await websocket.send_text(
                    json.dumps(
                        {
                            "type": "snapshot",
                            "workspace_id": latest.workspace_id,
                            "file_path": latest.file_path,
                            "version": latest.version,
                            "content": latest.content,
                            "read_only": latest.read_only,
                        }
                    )
                )
                continue
            await websocket.send_text(
                json.dumps(
                    {
                        "type": "ack",
                        "workspace_id": updated.workspace_id,
                        "file_path": updated.file_path,
                        "version": updated.version,
                    }
                )
            )

    except WebSocketDisconnect:
        await _cleanup_collab()
    except Exception:
        # Client disconnected during send (e.g. initial snapshot/presence).
        # Treat identically to WebSocketDisconnect for cleanup purposes.
        await _cleanup_collab()


@router.post("/collab/workspaces/{workspace_id}/files/create")
async def collab_create_file(
    workspace_id: str,
    payload: dict[str, Any],
    request: Request,
):
    token = _extract_capability_token_from_request(request)
    _, user_id = _require_workspace_capability(
        token,
        workspace_id,
        "userspace.collab_mutate",
    )
    file_path = str(payload.get("file_path", "")).strip()
    content = str(payload.get("content", ""))
    created = await _runtime_service().create_collab_file(
        workspace_id,
        file_path,
        user_id,
        content=content,
    )
    await _broadcast_collab_message(
        workspace_id,
        created.file_path,
        {
            "type": "file_created",
            "workspace_id": workspace_id,
            "file_path": created.file_path,
            "version": created.version,
        },
    )
    return {
        "success": True,
        "workspace_id": workspace_id,
        "file_path": created.file_path,
        "version": created.version,
    }


@router.post("/collab/workspaces/{workspace_id}/files/rename")
async def collab_rename_file(
    workspace_id: str,
    payload: dict[str, Any],
    request: Request,
):
    token = _extract_capability_token_from_request(request)
    _, user_id = _require_workspace_capability(
        token,
        workspace_id,
        "userspace.collab_mutate",
    )
    old_path = str(payload.get("old_path", "")).strip()
    new_path = str(payload.get("new_path", "")).strip()
    result = await _runtime_service().rename_collab_file(
        workspace_id,
        old_path,
        new_path,
        user_id,
    )
    await _broadcast_collab_message(
        workspace_id,
        result["new_path"],
        {
            "type": "file_renamed",
            "workspace_id": workspace_id,
            "old_path": result["old_path"],
            "new_path": result["new_path"],
        },
    )
    return result


@router.post("/collab/workspaces/{workspace_id}/files/delete")
async def collab_delete_file(
    workspace_id: str,
    payload: dict[str, Any],
    request: Request,
):
    token = _extract_capability_token_from_request(request)
    _, user_id = _require_workspace_capability(
        token,
        workspace_id,
        "userspace.collab_mutate",
    )
    file_path = str(payload.get("file_path", "")).strip()
    result = await _runtime_service().delete_collab_file(
        workspace_id,
        file_path,
        user_id,
    )
    return result


# ---------------------------------------------------------------------------
# Same-origin path-based preview proxy
#
# These routes remain available for backwards compatibility with existing
# preview URLs and explicit root-proxy flows, but preview-launch now prefers
# dedicated preview subdomains instead of implicitly falling back here.
# ---------------------------------------------------------------------------


@router.api_route("/workspaces/{workspace_id}/preview", methods=_PROXY_METHODS)
@router.api_route("/workspaces/{workspace_id}/preview/{path:path}", methods=_PROXY_METHODS)
async def workspace_preview_path_proxy(
    workspace_id: str,
    request: Request,
    path: str = "",
    user: Any = Depends(get_current_user),
) -> Response:
    await _userspace_service().enforce_workspace_role(workspace_id, user.id, "viewer")
    upstream_url = await _runtime_service().build_workspace_preview_upstream_url(
        workspace_id,
        user.id,
        path,
        query=_sanitize_preview_query(request.url.query),
    )
    return await _proxy_http_request(
        request,
        upstream_url,
        bridge_workspace_id=workspace_id,
    )


@router.websocket("/workspaces/{workspace_id}/preview")
@router.websocket("/workspaces/{workspace_id}/preview/{path:path}")
async def workspace_preview_path_websocket(
    workspace_id: str,
    websocket: WebSocket,
    path: str = "",
) -> None:
    session_token = websocket.cookies.get("ragtime_session", "").strip()
    token_data = decode_access_token(session_token) if session_token else None
    if not token_data:
        await websocket.close(code=4401)
        return
    try:
        await _userspace_service().enforce_workspace_role(workspace_id, token_data.user_id, "viewer")
    except HTTPException:
        await websocket.close(code=4403)
        return
    upstream_url = await _runtime_service().build_workspace_preview_upstream_url(
        workspace_id,
        token_data.user_id,
        path,
        query=_sanitize_preview_query(websocket.url.query),
    )
    await _proxy_websocket_request(websocket, _to_websocket_url(upstream_url))


@router.api_route("/shared/{share_token}/preview", methods=_PROXY_METHODS)
@router.api_route("/shared/{share_token}/preview/{path:path}", methods=_PROXY_METHODS)
async def shared_preview_path_proxy(
    share_token: str,
    request: Request,
    path: str = "",
    share_password: str | None = Header(default=None, alias="X-UserSpace-Share-Password"),
    user: Any | None = Depends(get_current_user_optional),
) -> Response:
    share_auth_token = share_auth_token_from_request(
        request.headers,
        request.cookies,
        share_token=share_token,
    )
    workspace_id = await _userspace_service().resolve_shared_workspace_id(
        share_token,
        current_user=user,
        password=share_password,
        share_auth_token=share_auth_token,
    )
    upstream_url = await _runtime_service().build_shared_preview_upstream_url(
        workspace_id,
        path,
        query=_sanitize_preview_query(request.url.query),
    )
    return await _proxy_http_request(
        request,
        upstream_url,
        bridge_workspace_id=workspace_id,
    )
